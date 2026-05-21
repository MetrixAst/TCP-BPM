import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


_HEADER_FONT  = Font(bold=True, color='FFFFFF')
_HEADER_FILL  = PatternFill(fill_type='solid', fgColor='1F3864')
_HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_ALT_FILL     = PatternFill(fill_type='solid', fgColor='EBF3FB')
_THIN_BORDER  = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _write_headers(ws, headers):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    ws.row_dimensions[1].height = 20


def _apply_row_style(ws, row_idx):
    fill = _ALT_FILL if row_idx % 2 == 0 else None
    for cell in ws[row_idx]:
        if fill:
            cell.fill = fill
        cell.border = _THIN_BORDER


def export_payment_registry(queryset) -> HttpResponse:
    """Экспорт TenantPaymentRegistry в Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Реестр оплат'

    headers = ['Арендатор', 'Договор', 'Период', 'Начислено', 'Оплачено', 'Задолженность', 'Статус']
    _write_headers(ws, headers)

    STATUS_LABELS = {
        'paid': 'Оплачен',
        'partial': 'Частично оплачен',
        'pending': 'Ожидает оплаты',
        'overdue': 'Просрочен',
        'cancelled': 'Отменён',
    }

    for row_idx, entry in enumerate(queryset, start=2):
        ws.append([
            str(entry.tenant),
            entry.contract_number,
            entry.period.strftime('%m.%Y') if entry.period else '',
            float(entry.charged),
            float(entry.paid),
            float(entry.balance),
            STATUS_LABELS.get(entry.status, entry.status),
        ])
        _apply_row_style(ws, row_idx)

    _auto_width(ws)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="payment_registry.xlsx"'
    wb.save(response)
    return response


def export_budget(queryset) -> HttpResponse:
    """Экспорт BudgetItem в Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Бюджет'

    headers = ['Категория', 'Тип периода', 'Год', 'Месяц', 'Квартал', 'План', 'Факт', 'Прогноз', 'Отклонение', 'Исполнение %']
    _write_headers(ws, headers)

    for row_idx, item in enumerate(queryset, start=2):
        ws.append([
            str(item.category),
            item.get_period_type_display(),
            item.year,
            item.month or '',
            item.quarter or '',
            float(item.plan),
            float(item.fact),
            float(item.forecast),
            float(item.variance),
            item.execution_pct or '',
        ])
        _apply_row_style(ws, row_idx)

    _auto_width(ws)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="budget.xlsx"'
    wb.save(response)
    return response


def export_cashflow(queryset) -> HttpResponse:
    """Экспорт CashFlowRecord в Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ДДС'

    headers = ['Дата', 'Направление', 'Тип', 'Сумма', 'Валюта', 'Контрагент', 'Назначение', 'Номер документа']
    _write_headers(ws, headers)

    DIR_LABELS  = {'inflow': 'Поступление', 'outflow': 'Списание'}
    TYPE_LABELS = {'operating': 'Операционная', 'investing': 'Инвестиционная', 'financing': 'Финансовая'}

    for row_idx, record in enumerate(queryset, start=2):
        ws.append([
            str(record.transaction_date),
            DIR_LABELS.get(record.direction, record.direction),
            TYPE_LABELS.get(record.flow_type, record.flow_type),
            float(record.amount),
            record.currency,
            str(record.counterparty) if record.counterparty else '',
            record.description or '',
            record.document_number or '',
        ])
        _apply_row_style(ws, row_idx)

    _auto_width(ws)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="cashflow.xlsx"'
    wb.save(response)
    return response


def export_financial_statement(queryset) -> HttpResponse:
    """Экспорт FinancialStatement в Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ОПиУ'

    headers = [
        'Период', 'Тип периода',
        'Выручка план', 'Выручка факт', 'Выручка прогноз',
        'EBITDA план', 'EBITDA факт',
        'Операционная прибыль план', 'Операционная прибыль факт',
        'Чистая прибыль план', 'Чистая прибыль факт',
    ]
    _write_headers(ws, headers)

    for row_idx, stmt in enumerate(queryset, start=2):
        ws.append([
            stmt.get_period_label(),
            stmt.get_period_type_display(),
            float(stmt.revenue_plan),
            float(stmt.revenue_fact),
            float(stmt.revenue_forecast),
            float(stmt.ebitda_plan),
            float(stmt.ebitda_fact),
            float(stmt.operating_profit_plan),
            float(stmt.operating_profit_fact),
            float(stmt.net_profit_plan),
            float(stmt.net_profit_fact),
        ])
        _apply_row_style(ws, row_idx)

    _auto_width(ws)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="financial_statement.xlsx"'
    wb.save(response)
    return response
