import io
import openpyxl
from django.test import TestCase, Client
from django.urls import reverse
from django.http import Http404
from datetime import date
from .models import LeaveRequest, LeaveType, WorkCalendar, Company, Position
from .enums import LeaveStatusEnum, DayTypeEnum
from account.role_permissions import RoleEnums
try:
    from account.models import Employee, UserAccount, Department
except ImportError:
    from apps.account.models import Employee, UserAccount, Department


class LeaveRequestLogicTest(TestCase):
    def setUp(self):
        self.company = Company.objects.create(name="TechCorp", bin_number="123456789012")
        self.dept = Department.objects.create(name="IT", company=self.company)
        self.user = UserAccount.objects.create(email="darya@test.kz", first_name="Dar")
        self.employee = Employee.objects.create(user=self.user, department=self.dept)

        self.leave_type = LeaveType.objects.create(
            name="Ежегодный", is_paid=True, max_days_per_year=24
        )

        WorkCalendar.objects.create(date=date(2026, 1, 1), day_type=DayTypeEnum.HOLIDAY, year=2026)
        WorkCalendar.objects.create(date=date(2026, 1, 2), day_type=DayTypeEnum.HOLIDAY, year=2026)

    def test_basic_working_days(self):
        leave = LeaveRequest.objects.create(
            employee=self.employee,
            leave_type=self.leave_type,
            start_date=date(2026, 4, 13),
            end_date=date(2026, 4, 17)
        )
        self.assertEqual(leave.working_days_count, 5)

    def test_holiday_exclusion(self):
        leave = LeaveRequest.objects.create(
            employee=self.employee,
            leave_type=self.leave_type,
            start_date=date(2025, 12, 31),
            end_date=date(2026, 1, 5)
        )
        self.assertEqual(leave.working_days_count, 2)

    def test_company_specific_override(self):
        special_date = date(2026, 4, 18)
        WorkCalendar.objects.create(
            date=special_date,
            day_type=DayTypeEnum.WORKING,
            company=self.company,
            year=2026
        )
        leave = LeaveRequest.objects.create(
            employee=self.employee,
            leave_type=self.leave_type,
            start_date=special_date,
            end_date=special_date
        )
        self.assertEqual(leave.working_days_count, 1)

    def test_status_workflow(self):
        leave = LeaveRequest.objects.create(
            employee=self.employee,
            leave_type=self.leave_type,
            start_date=date(2026, 6, 1),
            end_date=date(2026, 6, 5)
        )
        self.assertEqual(leave.status, LeaveStatusEnum.DRAFT)

        leave.status = LeaveStatusEnum.APPROVED
        leave.save()

        updated_leave = LeaveRequest.objects.get(pk=leave.pk)
        self.assertEqual(updated_leave.status, 'approved')

    def test_zero_days_on_weekend_only(self):
        leave = LeaveRequest.objects.create(
            employee=self.employee,
            leave_type=self.leave_type,
            start_date=date(2026, 4, 18),
            end_date=date(2026, 4, 19)
        )
        self.assertEqual(leave.working_days_count, 0)


class LeaveComplexViewsTest(TestCase):
    def setUp(self):
        self.company = Company.objects.create(name="DauInvest", bin_number="000000000000")
        self.dept = Department.objects.create(name="Analysis", company=self.company)

        self.u_emp = UserAccount.objects.create_user(username='darya', password='123')
        self.employee = Employee.objects.create(user=self.u_emp, department=self.dept)

        self.u_boss = UserAccount.objects.create_user(username='manager', password='123', head=True)
        self.boss = Employee.objects.create(user=self.u_boss, department=self.dept, head=True)

        self.paid_type = LeaveType.objects.create(name="Paid Vacation", max_days_per_year=24)
        self.client = Client()

    def test_invalid_date_range(self):
        self.client.login(username='darya', password='123')
        data = {
            'leave_type': self.paid_type.id,
            'start_date': '2026-07-10',
            'end_date': '2026-07-01'
        }
        response = self.client.post(reverse('hr:leave_create'), data)
        self.assertContains(response, "Дата начала не может быть позже даты окончания")

    def test_leave_list_filters(self):
        LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.paid_type,
            start_date=date(2026, 8, 1), end_date=date(2026, 8, 5),
            status=LeaveStatusEnum.APPROVED
        )
        self.client.login(username='manager', password='123')
        response = self.client.get(
            reverse('hr:leave_list'),
            {'status': LeaveStatusEnum.APPROVED}
        )
        self.assertEqual(len(response.context['leaves']), 1)

    def test_reject_workflow(self):
        leave = LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.paid_type,
            start_date=date(2026, 6, 1), end_date=date(2026, 6, 5),
            status=LeaveStatusEnum.PENDING
        )
        self.client.login(username='manager', password='123')
        self.client.post(reverse('hr:leave_reject', args=[leave.pk]))
        leave.refresh_from_db()
        self.assertEqual(leave.status, LeaveStatusEnum.REJECTED)

    def test_cancel_own_request(self):
        leave = LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.paid_type,
            start_date=date(2026, 6, 1), end_date=date(2026, 6, 5),
            status=LeaveStatusEnum.DRAFT
        )
        self.client.login(username='darya', password='123')
        self.client.post(reverse('hr:leave_cancel', args=[leave.pk]))
        self.assertEqual(LeaveRequest.objects.filter(pk=leave.pk).count(), 0)

    def test_ajax_with_holiday(self):
        WorkCalendar.objects.create(
            date=date(2026, 7, 1),
            day_type=DayTypeEnum.HOLIDAY,
            company=self.company
        )
        self.client.login(username='darya', password='123')
        response = self.client.get(reverse('hr:ajax_calculate_days'), {
            'start': '2026-07-01',
            'end': '2026-07-03'
        })
        self.assertEqual(response.json()['days'], 2)

    def test_cannot_cancel_approved_leave(self):
        leave = LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.paid_type,
            start_date=date(2026, 6, 1), end_date=date(2026, 6, 5),
            status=LeaveStatusEnum.APPROVED
        )
        self.client.login(username='darya', password='123')
        self.client.post(reverse('hr:leave_cancel', args=[leave.pk]))
        self.assertEqual(LeaveRequest.objects.filter(pk=leave.pk).count(), 1)



class LeaveListViewExtendedTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.company = Company.objects.create(name="ListCorp", bin_number="111111111111")
        self.dept = Department.objects.create(name="Dev", company=self.company)
        self.dept2 = Department.objects.create(name="QA", company=self.company)

        self.u_emp = UserAccount.objects.create_user(username='emp_list', password='123')
        self.employee = Employee.objects.create(user=self.u_emp, department=self.dept)

        self.u_emp2 = UserAccount.objects.create_user(username='emp_list2', password='123')
        self.employee2 = Employee.objects.create(user=self.u_emp2, department=self.dept2)

        self.u_boss = UserAccount.objects.create_user(username='boss_list', password='123')
        self.boss = Employee.objects.create(user=self.u_boss, department=self.dept, head=True)

        self.type1 = LeaveType.objects.create(name="Ежегодный_л", max_days_per_year=24)
        self.type2 = LeaveType.objects.create(name="Учебный_л", max_days_per_year=10)

    def test_anonymous_redirected(self):
        response = self.client.get(reverse('hr:leave_list'))
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response.url)

    def test_logged_in_sees_page(self):
        self.client.login(username='emp_list', password='123')
        response = self.client.get(reverse('hr:leave_list'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'site/hr/leave_list.html')

    def test_context_keys_present(self):
        self.client.login(username='emp_list', password='123')
        response = self.client.get(reverse('hr:leave_list'))
        self.assertIn('leaves', response.context)
        self.assertIn('filter_form', response.context)
        self.assertIn('is_manager', response.context)

    def test_is_manager_false_for_regular_employee(self):
        self.client.login(username='emp_list', password='123')
        response = self.client.get(reverse('hr:leave_list'))
        self.assertFalse(response.context['is_manager'])

    def test_is_manager_true_for_head(self):
        self.client.login(username='boss_list', password='123')
        response = self.client.get(reverse('hr:leave_list'))
        self.assertTrue(response.context['is_manager'])

    def test_all_leaves_shown_without_filters(self):
        LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.type1,
            start_date=date(2026, 8, 3), end_date=date(2026, 8, 7)
        )
        LeaveRequest.objects.create(
            employee=self.employee2, leave_type=self.type2,
            start_date=date(2026, 8, 10), end_date=date(2026, 8, 14)
        )
        self.client.login(username='emp_list', password='123')
        response = self.client.get(reverse('hr:leave_list'))
        self.assertEqual(response.context['leaves'].count(), 2)

    def test_filter_by_leave_type(self):
        LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.type1,
            start_date=date(2026, 8, 3), end_date=date(2026, 8, 7)
        )
        LeaveRequest.objects.create(
            employee=self.employee2, leave_type=self.type2,
            start_date=date(2026, 8, 10), end_date=date(2026, 8, 14)
        )
        self.client.login(username='emp_list', password='123')
        response = self.client.get(
            reverse('hr:leave_list'), {'leave_type': self.type1.pk}
        )
        leaves = list(response.context['leaves'])
        self.assertEqual(len(leaves), 1)
        self.assertEqual(leaves[0].leave_type, self.type1)

    def test_filter_by_department(self):
        LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.type1,
            start_date=date(2026, 8, 3), end_date=date(2026, 8, 7)
        )
        LeaveRequest.objects.create(
            employee=self.employee2, leave_type=self.type1,
            start_date=date(2026, 8, 10), end_date=date(2026, 8, 14)
        )
        self.client.login(username='emp_list', password='123')
        response = self.client.get(
            reverse('hr:leave_list'), {'department': self.dept.pk}
        )
        leaves = list(response.context['leaves'])
        self.assertEqual(len(leaves), 1)
        self.assertEqual(leaves[0].employee.department, self.dept)

    def test_filter_by_date_from(self):
        LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.type1,
            start_date=date(2026, 8, 1), end_date=date(2026, 8, 5)
        )
        LeaveRequest.objects.create(
            employee=self.employee2, leave_type=self.type1,
            start_date=date(2026, 9, 1), end_date=date(2026, 9, 5)
        )
        self.client.login(username='emp_list', password='123')
        response = self.client.get(
            reverse('hr:leave_list'), {'date_from': '2026-08-20'}
        )
        leaves = list(response.context['leaves'])
        self.assertEqual(len(leaves), 1)
        self.assertEqual(leaves[0].start_date, date(2026, 9, 1))

    def test_filter_by_date_to(self):
        LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.type1,
            start_date=date(2026, 8, 1), end_date=date(2026, 8, 5)
        )
        LeaveRequest.objects.create(
            employee=self.employee2, leave_type=self.type1,
            start_date=date(2026, 9, 1), end_date=date(2026, 9, 30)
        )
        self.client.login(username='emp_list', password='123')
        response = self.client.get(
            reverse('hr:leave_list'), {'date_to': '2026-08-10'}
        )
        leaves = list(response.context['leaves'])
        self.assertEqual(len(leaves), 1)
        self.assertEqual(leaves[0].end_date, date(2026, 8, 5))

    def test_filter_by_search_first_name(self):
        self.u_emp.first_name = 'Жанар'
        self.u_emp.save()
        LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.type1,
            start_date=date(2026, 8, 3), end_date=date(2026, 8, 7)
        )
        LeaveRequest.objects.create(
            employee=self.employee2, leave_type=self.type1,
            start_date=date(2026, 8, 10), end_date=date(2026, 8, 14)
        )
        self.client.login(username='emp_list', password='123')
        response = self.client.get(reverse('hr:leave_list'), {'search': 'Жанар'})
        leaves = list(response.context['leaves'])
        self.assertEqual(len(leaves), 1)
        self.assertEqual(leaves[0].employee, self.employee)

    def test_empty_list_when_no_leaves(self):
        self.client.login(username='emp_list', password='123')
        response = self.client.get(reverse('hr:leave_list'))
        self.assertEqual(response.context['leaves'].count(), 0)


class LeaveCreateExtendedTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.company = Company.objects.create(name="CreateCorp", bin_number="222222222222")
        self.dept = Department.objects.create(name="HR_dept", company=self.company)

        self.u_emp = UserAccount.objects.create_user(username='emp_create', password='123')
        self.employee = Employee.objects.create(user=self.u_emp, department=self.dept)

        self.u_no_profile = UserAccount.objects.create_user(username='noprofile', password='123')

        self.leave_type = LeaveType.objects.create(name="Ежегодный_c", max_days_per_year=24)

    def test_anonymous_redirected(self):
        response = self.client.get(reverse('hr:leave_create'))
        self.assertEqual(response.status_code, 302)

    def test_get_shows_form(self):
        self.client.login(username='emp_create', password='123')
        response = self.client.get(reverse('hr:leave_create'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'site/hr/leave_create.html')
        self.assertIn('form', response.context)

    def test_successful_create_redirects_to_list(self):
        self.client.login(username='emp_create', password='123')
        response = self.client.post(reverse('hr:leave_create'), {
            'leave_type': self.leave_type.pk,
            'start_date': '2026-08-03',
            'end_date': '2026-08-07',
            'comment': 'Тест',
        })
        self.assertRedirects(response, reverse('hr:leave_list'))

    def test_successful_create_saves_to_db(self):
        self.client.login(username='emp_create', password='123')
        self.client.post(reverse('hr:leave_create'), {
            'leave_type': self.leave_type.pk,
            'start_date': '2026-08-03',
            'end_date': '2026-08-07',
            'comment': '',
        })
        self.assertEqual(LeaveRequest.objects.count(), 1)

    def test_create_sets_status_pending(self):
        self.client.login(username='emp_create', password='123')
        self.client.post(reverse('hr:leave_create'), {
            'leave_type': self.leave_type.pk,
            'start_date': '2026-08-03',
            'end_date': '2026-08-07',
            'comment': '',
        })
        leave = LeaveRequest.objects.first()
        self.assertEqual(leave.status, LeaveStatusEnum.PENDING)

    def test_create_sets_correct_employee(self):
        self.client.login(username='emp_create', password='123')
        self.client.post(reverse('hr:leave_create'), {
            'leave_type': self.leave_type.pk,
            'start_date': '2026-08-03',
            'end_date': '2026-08-07',
            'comment': '',
        })
        leave = LeaveRequest.objects.first()
        self.assertEqual(leave.employee, self.employee)

    def test_user_without_profile_redirected_to_list(self):
        self.client.login(username='noprofile', password='123')
        response = self.client.post(reverse('hr:leave_create'), {
            'leave_type': self.leave_type.pk,
            'start_date': '2026-08-03',
            'end_date': '2026-08-07',
            'comment': '',
        })
        self.assertRedirects(response, reverse('hr:leave_list'))
        self.assertEqual(LeaveRequest.objects.count(), 0)

    def test_missing_leave_type_shows_form_with_error(self):
        self.client.login(username='emp_create', password='123')
        response = self.client.post(reverse('hr:leave_create'), {
            'leave_type': '',
            'start_date': '2026-08-03',
            'end_date': '2026-08-07',
        })
        self.assertEqual(LeaveRequest.objects.count(), 0)
        self.assertIn(response.status_code, [200, 302])


class LeaveDetailViewTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.company = Company.objects.create(name="DetailCorp", bin_number="333333333333")
        self.dept = Department.objects.create(name="Fin", company=self.company)

        self.u_owner = UserAccount.objects.create_user(username='owner_detail', password='123')
        self.u_owner.role = RoleEnums.ADMINISTRATOR.value
        self.u_owner.save()
        self.owner = Employee.objects.create(user=self.u_owner, department=self.dept)

        self.u_other = UserAccount.objects.create_user(username='other_detail', password='123')
        self.other = Employee.objects.create(user=self.u_other, department=self.dept)

        self.u_boss = UserAccount.objects.create_user(username='boss_detail', password='123')
        self.u_boss.role = RoleEnums.ADMINISTRATOR.value
        self.u_boss.save()
        self.boss = Employee.objects.create(user=self.u_boss, department=self.dept, head=True)

        self.leave_type = LeaveType.objects.create(name="Ежегодный_d", max_days_per_year=24)
        self.leave = LeaveRequest.objects.create(
            employee=self.owner, leave_type=self.leave_type,
            start_date=date(2026, 8, 3), end_date=date(2026, 8, 7),
            status=LeaveStatusEnum.PENDING,
        )

    def get_url(self):
        return reverse('hr:leave_detail', kwargs={'pk': self.leave.pk})

    def test_anonymous_redirected(self):
        response = self.client.get(self.get_url())
        self.assertEqual(response.status_code, 302)

    def test_owner_can_view(self):
        self.client.login(username='owner_detail', password='123')
        response = self.client.get(self.get_url())
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'site/hr/leave_detail.html')

    def test_manager_can_view(self):
        self.client.login(username='boss_detail', password='123')
        response = self.client.get(self.get_url())
        self.assertEqual(response.status_code, 200)

    def test_other_employee_redirected(self):
        self.client.login(username='other_detail', password='123')
        response = self.client.get(self.get_url())
        self.assertRedirects(response, reverse('hr:leave_list'))

    def test_context_has_correct_leave(self):
        self.client.login(username='owner_detail', password='123')
        response = self.client.get(self.get_url())
        self.assertEqual(response.context['leave'].pk, self.leave.pk)

    def test_is_owner_true_for_owner(self):
        self.client.login(username='owner_detail', password='123')
        response = self.client.get(self.get_url())
        self.assertTrue(response.context['is_owner'])

    def test_is_owner_false_for_manager(self):
        self.client.login(username='boss_detail', password='123')
        response = self.client.get(self.get_url())
        self.assertFalse(response.context['is_owner'])

    def test_nonexistent_returns_404(self):
        self.client.login(username='owner_detail', password='123')
        
        from django.test import RequestFactory
        factory = RequestFactory()
        request = factory.get(reverse('hr:leave_detail', kwargs={'pk': 99999}))
        
        request.user = self.u_owner 
        from .views import leave_detail
        with self.assertRaises(Http404):
            leave_detail(request, pk=99999)


class LeaveApproveViewTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.company = Company.objects.create(name="ApproveCorp", bin_number="444444444444")
        self.dept = Department.objects.create(name="Ops", company=self.company)

        self.u_emp = UserAccount.objects.create_user(username='emp_approve', password='123')
        self.employee = Employee.objects.create(user=self.u_emp, department=self.dept)

        self.u_emp2 = UserAccount.objects.create_user(username='emp_approve2', password='123')
        self.employee2 = Employee.objects.create(user=self.u_emp2, department=self.dept)

        self.u_boss = UserAccount.objects.create_user(username='boss_approve', password='123')
        self.u_boss.role = RoleEnums.ADMINISTRATOR.value
        self.u_boss.save()
        self.boss = Employee.objects.create(user=self.u_boss, department=self.dept, head=True)

        self.leave_type = LeaveType.objects.create(name="Ежегодный_a", max_days_per_year=24)
        self.leave = LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.leave_type,
            start_date=date(2026, 8, 3), end_date=date(2026, 8, 7),
            status=LeaveStatusEnum.PENDING,
        )

    def get_url(self):
        return reverse('hr:leave_approve', kwargs={'pk': self.leave.pk})

    def test_anonymous_redirected(self):
        response = self.client.post(self.get_url())
        self.assertEqual(response.status_code, 302)

    def test_get_request_does_not_approve(self):
        self.client.login(username='boss_approve', password='123')
        self.client.get(self.get_url())
        self.leave.refresh_from_db()
        self.assertEqual(self.leave.status, LeaveStatusEnum.PENDING)

    def test_manager_can_approve_via_post(self):
        self.client.login(username='boss_approve', password='123')
        response = self.client.post(self.get_url())
        self.assertRedirects(response, reverse('hr:leave_list'))
        self.leave.refresh_from_db()
        self.assertEqual(self.leave.status, LeaveStatusEnum.APPROVED)

    def test_approve_sets_approver(self):
        self.client.login(username='boss_approve', password='123')
        self.client.post(self.get_url())
        self.leave.refresh_from_db()
        self.assertEqual(self.leave.approver, self.boss)

    def test_regular_employee_cannot_approve(self):
        self.client.login(username='emp_approve2', password='123')
        self.client.post(self.get_url())
        self.leave.refresh_from_db()
        self.assertEqual(self.leave.status, LeaveStatusEnum.PENDING)

    def test_cannot_approve_already_approved(self):
        self.leave.status = LeaveStatusEnum.APPROVED
        self.leave.save()
        self.client.login(username='boss_approve', password='123')
        self.client.post(self.get_url())
        self.leave.refresh_from_db()
        self.assertEqual(self.leave.status, LeaveStatusEnum.APPROVED)

    def test_cannot_approve_rejected(self):
        self.leave.status = LeaveStatusEnum.REJECTED
        self.leave.save()
        self.client.login(username='boss_approve', password='123')
        self.client.post(self.get_url())
        self.leave.refresh_from_db()
        self.assertEqual(self.leave.status, LeaveStatusEnum.REJECTED)

    def test_nonexistent_returns_404(self):
        self.client.login(username='boss_approve', password='123')
        from .views import leave_approve
        with self.assertRaises(Http404):
            leave_approve(self.client.get('/').wsgi_request, pk=99999)


class LeaveRejectExtendedTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.company = Company.objects.create(name="RejectCorp", bin_number="555555555555")
        self.dept = Department.objects.create(name="Legal", company=self.company)

        self.u_emp = UserAccount.objects.create_user(username='emp_reject', password='123')
        self.employee = Employee.objects.create(user=self.u_emp, department=self.dept)

        self.u_emp2 = UserAccount.objects.create_user(username='emp_reject2', password='123')
        self.employee2 = Employee.objects.create(user=self.u_emp2, department=self.dept)

        self.u_boss = UserAccount.objects.create_user(username='boss_reject', password='123')
        self.u_boss.role = RoleEnums.ADMINISTRATOR.value
        self.u_boss.save()
        self.boss = Employee.objects.create(user=self.u_boss, department=self.dept, head=True)

        self.leave_type = LeaveType.objects.create(name="Ежегодный_r", max_days_per_year=24)
        self.leave = LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.leave_type,
            start_date=date(2026, 8, 3), end_date=date(2026, 8, 7),
            status=LeaveStatusEnum.PENDING,
        )

    def get_url(self):
        return reverse('hr:leave_reject', kwargs={'pk': self.leave.pk})

    def test_get_request_does_not_reject(self):
        self.client.login(username='boss_reject', password='123')
        self.client.get(self.get_url())
        self.leave.refresh_from_db()
        self.assertEqual(self.leave.status, LeaveStatusEnum.PENDING)

    def test_regular_employee_cannot_reject(self):
        self.client.login(username='emp_reject2', password='123')
        self.client.post(self.get_url())
        self.leave.refresh_from_db()
        self.assertEqual(self.leave.status, LeaveStatusEnum.PENDING)

    def test_cannot_reject_already_approved(self):
        self.leave.status = LeaveStatusEnum.APPROVED
        self.leave.save()
        self.client.login(username='boss_reject', password='123')
        self.client.post(self.get_url())
        self.leave.refresh_from_db()
        self.assertEqual(self.leave.status, LeaveStatusEnum.APPROVED)

    def test_nonexistent_returns_404(self):
        self.client.login(username='boss_reject', password='123')
        from .views import leave_reject
        with self.assertRaises(Http404):
            leave_reject(self.client.get('/').wsgi_request, pk=99999)


class LeaveCancelExtendedTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.company = Company.objects.create(name="CancelCorp", bin_number="666666666666")
        self.dept = Department.objects.create(name="Mkt", company=self.company)

        self.u_emp = UserAccount.objects.create_user(username='emp_cancel', password='123')
        self.employee = Employee.objects.create(user=self.u_emp, department=self.dept)

        self.u_emp2 = UserAccount.objects.create_user(username='emp_cancel2', password='123')
        self.employee2 = Employee.objects.create(user=self.u_emp2, department=self.dept)

        self.leave_type = LeaveType.objects.create(name="Ежегодный_cn", max_days_per_year=24)

    def test_get_request_does_not_delete(self):
        leave = LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.leave_type,
            start_date=date(2026, 8, 3), end_date=date(2026, 8, 7),
            status=LeaveStatusEnum.PENDING,
        )
        self.client.login(username='emp_cancel', password='123')
        self.client.get(reverse('hr:leave_cancel', args=[leave.pk]))
        self.assertTrue(LeaveRequest.objects.filter(pk=leave.pk).exists())

    def test_can_cancel_pending_via_post(self):
        leave = LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.leave_type,
            start_date=date(2026, 8, 3), end_date=date(2026, 8, 7),
            status=LeaveStatusEnum.PENDING,
        )
        self.client.login(username='emp_cancel', password='123')
        self.client.post(reverse('hr:leave_cancel', args=[leave.pk]))
        self.assertFalse(LeaveRequest.objects.filter(pk=leave.pk).exists())

    def test_other_employee_cannot_cancel(self):
        leave = LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.leave_type,
            start_date=date(2026, 8, 3), end_date=date(2026, 8, 7),
            status=LeaveStatusEnum.PENDING,
        )
        self.client.login(username='emp_cancel2', password='123')
        self.client.post(reverse('hr:leave_cancel', args=[leave.pk]))
        self.assertTrue(LeaveRequest.objects.filter(pk=leave.pk).exists())

    def test_anonymous_redirected(self):
        leave = LeaveRequest.objects.create(
            employee=self.employee, leave_type=self.leave_type,
            start_date=date(2026, 8, 3), end_date=date(2026, 8, 7),
            status=LeaveStatusEnum.PENDING,
        )
        response = self.client.post(reverse('hr:leave_cancel', args=[leave.pk]))
        self.assertEqual(response.status_code, 302)
        self.assertTrue(LeaveRequest.objects.filter(pk=leave.pk).exists())


class AjaxCalculateDaysExtendedTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.company = Company.objects.create(name="AjaxCorp", bin_number="777777777777")
        self.dept = Department.objects.create(name="Data", company=self.company)

        self.u_emp = UserAccount.objects.create_user(username='emp_ajax', password='123')
        self.employee = Employee.objects.create(user=self.u_emp, department=self.dept)

        self.u_no_profile = UserAccount.objects.create_user(username='noprofile_ajax', password='123')

    def get_url(self):
        return reverse('hr:ajax_calculate_days')

    def test_anonymous_redirected(self):
        response = self.client.get(self.get_url(), {'start': '2026-06-01', 'end': '2026-06-05'})
        self.assertEqual(response.status_code, 302)

    def test_returns_json_content_type(self):
        self.client.login(username='emp_ajax', password='123')
        response = self.client.get(self.get_url(), {'start': '2026-06-01', 'end': '2026-06-05'})
        self.assertEqual(response.status_code, 200)
        self.assertIn('application/json', response['Content-Type'])

    def test_monday_to_friday_five_days(self):
        """2026-06-01 (пн) — 2026-06-05 (пт) = 5 рабочих дней."""
        self.client.login(username='emp_ajax', password='123')
        response = self.client.get(self.get_url(), {'start': '2026-06-01', 'end': '2026-06-05'})
        self.assertEqual(response.json()['days'], 5)

    def test_weekend_excluded(self):
        """2026-06-01 (пн) — 2026-06-07 (вс) = 5 рабочих дней."""
        self.client.login(username='emp_ajax', password='123')
        response = self.client.get(self.get_url(), {'start': '2026-06-01', 'end': '2026-06-07'})
        self.assertEqual(response.json()['days'], 5)

    def test_no_params_returns_zero(self):
        self.client.login(username='emp_ajax', password='123')
        response = self.client.get(self.get_url())
        self.assertEqual(response.json()['days'], 0)

    def test_missing_end_returns_zero(self):
        self.client.login(username='emp_ajax', password='123')
        response = self.client.get(self.get_url(), {'start': '2026-06-01'})
        self.assertEqual(response.json()['days'], 0)

    def test_invalid_format_returns_zero(self):
        self.client.login(username='emp_ajax', password='123')
        response = self.client.get(self.get_url(), {'start': 'abc', 'end': 'def'})
        self.assertEqual(response.json()['days'], 0)

    def test_start_after_end_returns_zero(self):
        self.client.login(username='emp_ajax', password='123')
        response = self.client.get(self.get_url(), {'start': '2026-06-10', 'end': '2026-06-05'})
        self.assertEqual(response.json()['days'], 0)

    def test_user_without_employee_profile_returns_zero(self):
        self.client.login(username='noprofile_ajax', password='123')
        response = self.client.get(self.get_url(), {'start': '2026-06-01', 'end': '2026-06-05'})
        self.assertEqual(response.json()['days'], 0)

    def test_single_working_day(self):
        self.client.login(username='emp_ajax', password='123')
        response = self.client.get(self.get_url(), {'start': '2026-06-01', 'end': '2026-06-01'})
        self.assertEqual(response.json()['days'], 1)

    def test_single_weekend_returns_zero(self):
        self.client.login(username='emp_ajax', password='123')
        response = self.client.get(self.get_url(), {'start': '2026-06-06', 'end': '2026-06-06'})
        self.assertEqual(response.json()['days'], 0)


class LeaveFormTest(TestCase):

    def setUp(self):
        self.leave_type = LeaveType.objects.create(name="Ежегодный_f", max_days_per_year=24)

    def test_leave_request_form_valid(self):
        from .forms import LeaveRequestForm
        form = LeaveRequestForm(data={
            'leave_type': self.leave_type.pk,
            'start_date': '2026-08-03',
            'end_date': '2026-08-07',
            'comment': '',
        })
        self.assertTrue(form.is_valid(), form.errors)

    def test_leave_request_form_start_after_end(self):
        from .forms import LeaveRequestForm
        form = LeaveRequestForm(data={
            'leave_type': self.leave_type.pk,
            'start_date': '2026-08-10',
            'end_date': '2026-08-05',
            'comment': '',
        })
        self.assertFalse(form.is_valid())
        self.assertIn('__all__', form.errors)

    def test_leave_request_form_equal_dates_valid(self):
        from .forms import LeaveRequestForm
        form = LeaveRequestForm(data={
            'leave_type': self.leave_type.pk,
            'start_date': '2026-08-03',
            'end_date': '2026-08-03',
            'comment': '',
        })
        self.assertTrue(form.is_valid(), form.errors)

    def test_leave_request_form_missing_type(self):
        from .forms import LeaveRequestForm
        form = LeaveRequestForm(data={
            'leave_type': '',
            'start_date': '2026-08-03',
            'end_date': '2026-08-07',
            'comment': '',
        })
        if form.is_valid():
            pass
        else:
            self.assertIn('leave_type', form.errors)

    def test_leave_filter_form_empty_is_valid(self):
        from .forms import LeaveFilterForm
        form = LeaveFilterForm(data={})
        self.assertTrue(form.is_valid(), form.errors)

    def test_leave_filter_form_invalid_date(self):
        from .forms import LeaveFilterForm
        form = LeaveFilterForm(data={'date_from': 'notadate'})
        self.assertFalse(form.is_valid())

class LeaveTimelineAndExportTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.company = Company.objects.create(name="ExportCorp", bin_number="888888888888")
        self.dept1 = Department.objects.create(name="Analytics", company=self.company)
        self.dept2 = Department.objects.create(name="Support", company=self.company)

        self.u_emp1 = UserAccount.objects.create_user(username='emp_export1', password='123')
        self.employee1 = Employee.objects.create(user=self.u_emp1, department=self.dept1)

        self.u_emp2 = UserAccount.objects.create_user(username='emp_export2', password='123')
        self.employee2 = Employee.objects.create(user=self.u_emp2, department=self.dept2)

        self.leave_type = LeaveType.objects.create(name="Ежегодный_exp", max_days_per_year=24)

        self.leave1 = LeaveRequest.objects.create(
            employee=self.employee1, leave_type=self.leave_type,
            start_date=date(2026, 8, 1), end_date=date(2026, 8, 10),
            status=LeaveStatusEnum.APPROVED,
        )
        self.leave2 = LeaveRequest.objects.create(
            employee=self.employee2, leave_type=self.leave_type,
            start_date=date(2026, 9, 1), end_date=date(2026, 9, 10),
            status=LeaveStatusEnum.PENDING,
        )

    def test_timeline_anonymous_redirected(self):
        response = self.client.get(reverse('hr:leave_timeline'))
        self.assertEqual(response.status_code, 302)

    def test_timeline_returns_json(self):
        self.client.login(username='emp_export1', password='123')
        response = self.client.get(reverse('hr:leave_timeline'))
        
        self.assertEqual(response.status_code, 200)
        self.assertIn('application/json', response.headers.get('Content-Type', ''))

        data = response.json()
        self.assertEqual(len(data), 2)
        self.assertEqual(data[0]['id'], self.leave1.id)
        self.assertIn('content', data[0])
        self.assertIn('group', data[0])
        self.assertIn('className', data[0])

    def test_timeline_filters_by_department(self):
        self.client.login(username='emp_export1', password='123')
        response = self.client.get(reverse('hr:leave_timeline'), {'department': self.dept1.pk})
        data = response.json()
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]['id'], self.leave1.id)

    def test_export_anonymous_redirected(self):
        response = self.client.get(reverse('hr:leave_export_excel'))
        self.assertEqual(response.status_code, 302)

    def test_export_returns_excel_headers(self):
        self.client.login(username='emp_export1', password='123')
        response = self.client.get(reverse('hr:leave_export_excel'))
        
        self.assertEqual(response.status_code, 200)
        self.assertIn('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', response.headers.get('Content-Type', ''))
        self.assertIn('attachment; filename="leaves_export.xlsx"', response.headers.get('Content-Disposition', ''))

    def test_export_excel_valid_content(self):
        self.client.login(username='emp_export1', password='123')
        response = self.client.get(reverse('hr:leave_export_excel'))
        
        wb = openpyxl.load_workbook(filename=io.BytesIO(response.content))
        ws = wb.active
        
        self.assertEqual(ws.title, "Отпуска")
        
        headers = [cell.value for cell in ws[1]]
        self.assertIn("ФИО сотрудника", headers)
        self.assertIn("Тип отпуска", headers)
        self.assertIn("Статус", headers)
        
        self.assertEqual(ws.max_row, 3)

    def test_export_filters_by_date(self):
        self.client.login(username='emp_export1', password='123')
        response = self.client.get(reverse('hr:leave_export_excel'), {'start_date': '2026-09-01'})
        wb = openpyxl.load_workbook(filename=io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.max_row, 2)