from django.test import TestCase, override_settings, Client
from django.utils import timezone
from datetime import date, timedelta, datetime

from account.models import UserAccount, Employee, Department
from hr.models import AttendanceRecord, Company
from hr.enums import CheckInEnum


def make_setup():
    company = Company.objects.create(name='Test Co BE12', bin_number='999888777001')
    dept = Department.objects.create(name='Test Dept BE12', company=company)
    user = UserAccount.objects.create_user(username='test_be12', password='pass', role='administrator')
    employee = Employee.objects.create(user=user, department=dept)
    return employee, user


class WorkTimeFormulaTest(TestCase):

    def setUp(self):
        self.employee, _ = make_setup()

    @override_settings(ATTENDANCE_DISABLED_TYPES=['lunch_start', 'lunch_end'])
    def test_formula_without_lunch_when_disabled(self):
        target = date(2026, 1, 5)
        base = timezone.make_aware(datetime(2026, 1, 5, 9, 0))
        AttendanceRecord.objects.bulk_create([
            AttendanceRecord(employee=self.employee, event_type=CheckInEnum.DAY_START, timestamp=base),
            AttendanceRecord(employee=self.employee, event_type=CheckInEnum.DAY_END, timestamp=base + timedelta(hours=9)),
        ])
        summary = AttendanceRecord.get_daily_summary(self.employee, target)
        self.assertEqual(summary['total_work_time'], timedelta(hours=9))
        self.assertFalse(summary['lunch_tracked'])

    @override_settings(ATTENDANCE_DISABLED_TYPES=[])
    def test_formula_with_lunch_when_enabled(self):
        target = date(2026, 1, 6)
        base = timezone.make_aware(datetime(2026, 1, 6, 9, 0))
        AttendanceRecord.objects.bulk_create([
            AttendanceRecord(employee=self.employee, event_type=CheckInEnum.DAY_START, timestamp=base),
            AttendanceRecord(employee=self.employee, event_type=CheckInEnum.LUNCH_START, timestamp=base + timedelta(hours=4)),
            AttendanceRecord(employee=self.employee, event_type=CheckInEnum.LUNCH_END, timestamp=base + timedelta(hours=5)),
            AttendanceRecord(employee=self.employee, event_type=CheckInEnum.DAY_END, timestamp=base + timedelta(hours=9)),
        ])
        summary = AttendanceRecord.get_daily_summary(self.employee, target)
        self.assertEqual(summary['total_work_time'], timedelta(hours=8))
        self.assertTrue(summary['lunch_tracked'])

    @override_settings(ATTENDANCE_DISABLED_TYPES=['lunch_start', 'lunch_end'])
    def test_old_lunch_records_not_recalculated(self):
        target = date(2026, 1, 7)
        base = timezone.make_aware(datetime(2026, 1, 7, 9, 0))
        AttendanceRecord.objects.bulk_create([
            AttendanceRecord(employee=self.employee, event_type=CheckInEnum.DAY_START, timestamp=base),
            AttendanceRecord(employee=self.employee, event_type=CheckInEnum.LUNCH_START, timestamp=base + timedelta(hours=4)),
            AttendanceRecord(employee=self.employee, event_type=CheckInEnum.LUNCH_END, timestamp=base + timedelta(hours=5)),
            AttendanceRecord(employee=self.employee, event_type=CheckInEnum.DAY_END, timestamp=base + timedelta(hours=9)),
        ])
        summary = AttendanceRecord.get_daily_summary(self.employee, target)
        self.assertEqual(summary['total_work_time'], timedelta(hours=9))
        self.assertFalse(summary['lunch_tracked'])


class AttendanceExportTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.employee, self.user = make_setup()
        self.client.login(username='test_be12', password='pass')

    @override_settings(ATTENDANCE_DISABLED_TYPES=['lunch_start', 'lunch_end'])
    def test_export_without_lunch_columns(self):
        from django.urls import reverse
        r = self.client.get(reverse('hr:attendance_export'))
        self.assertEqual(r.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            r['Content-Type']
        )
        import openpyxl, io
        wb = openpyxl.load_workbook(filename=io.BytesIO(r.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertNotIn('Начало обеда', headers)
        self.assertNotIn('Конец обеда', headers)
        self.assertIn('Приход', headers)
        self.assertIn('Уход', headers)

    @override_settings(ATTENDANCE_DISABLED_TYPES=[])
    def test_export_with_lunch_columns(self):
        from django.urls import reverse
        r = self.client.get(reverse('hr:attendance_export'))
        self.assertEqual(r.status_code, 200)
        import openpyxl, io
        wb = openpyxl.load_workbook(filename=io.BytesIO(r.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn('Начало обеда', headers)
        self.assertIn('Конец обеда', headers)