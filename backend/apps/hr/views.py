import openpyxl
import json
import base64
import secrets
from django.core.files.base import ContentFile
from django.core.exceptions import ValidationError, PermissionDenied
from django.shortcuts import redirect, render, get_object_or_404
from django.urls import reverse
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_date
from datetime import datetime, date, timedelta, time
from decimal import Decimal, ROUND_HALF_UP


from project.utils import get_or_none, get_or_error
from project.paginator import CustomPaginator, page_from_request

from account.role_permissions import need_permission, PermissionEnums, RolePermissions, RoleEnums
from account.models import Employee, Department
from account.forms import EmployeeForm

from .forms import (
    CalendarItemForm, EmployeeCreationForm, EmployeesListForm,
    LeaveFilterForm, LeaveRequestForm,
    EmployeeDocumentForm, DocumentFilterForm,
    EmployeeWorkPermitForm, PermitFilterForm,
    EmployeeCertificationForm, CertificationFilterForm,
    WorkScheduleForm,
)

from .models import (
    CalendarItem, Company, Position, LeaveRequest, 
    LeaveType, Vacation, SickLeave, EmploymentContract, AttendanceRecord, CheckInEnum, EmployeeDocument, EmployeeWorkPermit, EmployeeCertification, WorkCategory, WorkCalendar, EmployeeWorkSchedule
)
from .serializers import CalendarItemSerializer
from . import work_schedule as work_schedule_helper

import calendar as calendar_module

from .enums import CalendarItemType, LeaveStatusEnum, CertificationStatusEnum
from .access import (
    need_hr_directory,
    need_hr_registry,
    get_registry_access,
    can_view_employee,
    filter_by_access,
)
from .services import create_attendance_checkin

from esigner.services import send_for_signing


@need_hr_directory
def structure(request):
    role = request.user.role
    if hasattr(role, 'value'):
        role = role.value
    return render(request, 'site/hr/org.html', {
        'can_edit_org': RolePermissions.checkPermission(role, PermissionEnums.HR_COMPANIES),
        'departments_api_url': '/api/v1/hr/departments/',
        'companies_api_url': '/api/v1/hr/companies/',
    })


@need_hr_directory
def employees(request):

    page = 1

    queryset = Employee.objects.all()
    is_hr, is_head, curr = get_registry_access(request.user)
    if is_head and not is_hr and curr:
        queryset = queryset.filter(department=curr.department)
    ordered = False
    form = EmployeesListForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        filters = form.cleaned_data

        page = filters.get('page', 1)

        search = filters.get('search', '')
        if search != '':
            queryset = queryset.filter(Q(user__first_name__icontains=search) | Q(user__username__icontains=search))

        department = filters.get('department', None)
        if department is not None:
            queryset = queryset.filter(department=department)

        position = filters.get('position', None)
        if position is not None:
            queryset = queryset.filter(position=position)

        status = filters.get('status', '')
        if status != '':
            queryset = queryset.filter(status=status)

        ordering = filters.get('ordering', '')
        if ordering != '':
            if ordering == 'name':
                queryset = queryset.order_by('user__first_name')
                ordered = True
            elif ordering == 'department':
                queryset = queryset.order_by('department__name')
                ordered = True
            elif ordering == 'id':
                queryset = queryset.order_by('-id')
                ordered = True

    if not ordered:
        queryset = queryset.order_by('user__username')

    paginator = CustomPaginator(queryset, page)

    context = {
        'form': form,
        'paginator': paginator,
    }

    return render(request, 'site/hr/employees.html', context)


@need_hr_directory
@transaction.atomic
def create_employee(request):
    is_hr, _, _ = get_registry_access(request.user)
    if not is_hr:
        return HttpResponseForbidden()

    form = EmployeeCreationForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            user = form['user'].save()
            employee = form['employee'].save(commit=False)
            employee.user = user
            employee.save()

            if employee.head:
                employee.set_head()

            messages.success(
                request,
                'Сотрудник создан. Откройте оргструктуру — он уже отображается в дереве.',
            )
            return redirect(f"{reverse('hr:org')}?highlight=emp_{employee.pk}")

    context = {
        'form': form,
        'positions_by_department_url': reverse('hr:positions_by_department'),
    }

    return render(request, 'site/hr/create_employee.html', context)


def _build_employee_profile_context(request, employee, tab=None):
    import pytz
    LOCAL_TZ = pytz.timezone('Asia/Almaty')

    valid_tabs = ('overview', 'documents', 'certifications', 'permits', 'leaves', 'attendance')
    if tab is None:
        tab = request.GET.get('tab', 'overview')
    if tab not in valid_tabs:
        tab = 'overview'

    user = employee.user
    curr_employee = getattr(request.user, 'employee_info', None)
    is_own_profile = curr_employee is not None and curr_employee.pk == employee.pk
    is_hr, is_head, _ = get_registry_access(request.user)
    can_manage = (is_hr or is_head) and not is_own_profile

    documents = employee.documents.all().order_by('-created_at')
    certifications = employee.certifications.order_by('-issue_date')
    permits = employee.work_permits.select_related('category').order_by('expiry_date')
    leaves = employee.leave_requests.select_related('leave_type', 'approver__user').order_by('-start_date')

    from django.contrib.contenttypes.models import ContentType
    from esigner.models import ESignerSigning

    doc_ct = ContentType.objects.get_for_model(EmployeeDocument)
    permit_ct = ContentType.objects.get_for_model(EmployeeWorkPermit)

    doc_signings = {
        s.object_id: s for s in ESignerSigning.objects.filter(
            content_type=doc_ct, object_id__in=documents.values_list('pk', flat=True)
        )
    }
    permit_signings = {
        s.object_id: s for s in ESignerSigning.objects.filter(
            content_type=permit_ct, object_id__in=permits.values_list('pk', flat=True)
        )
    }

    today = date.today()
    month_start = today.replace(day=1)
    days_range = 30 if tab == 'attendance' else 7
    attendance_rows = []
    for offset in range(days_range):
        d = today - timedelta(days=offset)
        summary = AttendanceRecord.get_daily_summary(employee, d)
        events = summary.get('details', {})
        start_dt = events.get(CheckInEnum.DAY_START)
        end_dt = events.get(CheckInEnum.DAY_END)
        if start_dt:
            start_dt = start_dt.astimezone(LOCAL_TZ)
        if end_dt:
            end_dt = end_dt.astimezone(LOCAL_TZ)
        total_h = summary.get('total_work_time', timedelta(0)).total_seconds() / 3600
        attendance_rows.append({
            'date': d,
            'day_start': start_dt,
            'day_end': end_dt,
            'total_hours': total_h,
            'no_record': len(events) == 0,
        })

    attendance_month = (
        AttendanceRecord.objects.filter(
            employee=employee,
            event_type=CheckInEnum.DAY_START,
            timestamp__date__gte=month_start,
            timestamp__date__lte=today,
        )
        .values('timestamp__date')
        .distinct()
        .count()
    )

    subordinates = []
    if not is_own_profile:
        subordinates = list(
            Employee.objects.filter(supervisor=employee)
            .select_related('user', 'position')
            .order_by('user__last_name')[:12]
        )

    profile_base_url = reverse('hr:my_profile') if is_own_profile else reverse('hr:employee_detail', args=[employee.pk])

    return {
        'employee': employee,
        'user': user,
        'tab': tab,
        'tabs': valid_tabs,
        'can_manage': can_manage,
        'is_hr': is_hr, 
        'is_own_profile': is_own_profile,
        'profile_base_url': profile_base_url,
        'documents': documents,
        'certifications': certifications,
        'permits': permits,
        'doc_signings': doc_signings,
        'permit_signings': permit_signings,
        'leaves': leaves,
        'attendance_rows': attendance_rows,
        'subordinates': subordinates,
        'work_schedule': work_schedule_helper.get_schedule(employee),
        'stats': {
            'documents': documents.count(),
            'certifications': certifications.count(),
            'permits': permits.count(),
            'leaves': leaves.count(),
            'leaves_pending': leaves.filter(status=LeaveStatusEnum.PENDING).count(),
            'attendance_month': attendance_month,
        },
        'date_today': today,
    }


@need_permission(PermissionEnums.PROFILE)
def my_profile(request):
    employee = getattr(request.user, 'employee_info', None)
    if not employee:
        return render(request, 'site/hr/my_profile_empty.html')

    employee = (
        Employee.objects.select_related(
            'user', 'department', 'department__company', 'position', 'supervisor__user',
        )
        .filter(pk=employee.pk)
        .first()
    )
    context = _build_employee_profile_context(request, employee)
    return render(request, 'site/hr/employee_detail.html', context)


@need_permission(PermissionEnums.HR_SELF)
def employee_detail(request, pk):
    employee = get_object_or_404(
        Employee.objects.select_related(
            'user', 'department', 'department__company', 'position', 'supervisor__user',
        ),
        pk=pk,
    )
    if not can_view_employee(request.user, employee):
        return HttpResponseForbidden()
    context = _build_employee_profile_context(request, employee)
    return render(request, 'site/hr/employee_detail.html', context)


@need_hr_directory
def edit_employee(request, pk):
    is_hr, _, _ = get_registry_access(request.user)
    if not is_hr:
        return HttpResponseForbidden()

    employee = get_object_or_404(Employee, pk=pk)

    form = EmployeeForm(request.POST or None, instance=employee)

    if request.method == 'POST':
        if form.is_valid():

            employee = form.save()
            if employee.head:
                employee.set_head()

            messages.success(request, 'Данные сотрудника сохранены.')
            return redirect('hr:employee_detail', pk=pk)

    schedule = work_schedule_helper.get_schedule(employee)
    context = {
        'form': form,
        'employee': employee,
        'schedule_form': WorkScheduleForm(instance=schedule),
        'positions_by_department_url': reverse('hr:positions_by_department'),
    }

    return render(request, 'site/hr/edit_employee.html', context)


@need_hr_directory
def employee_schedule(request, pk):
    is_hr, _, _ = get_registry_access(request.user)
    if not is_hr:
        return HttpResponseForbidden()

    employee = get_object_or_404(Employee, pk=pk)
    schedule = work_schedule_helper.get_schedule(employee)

    if request.method == 'POST':
        form = WorkScheduleForm(request.POST, instance=schedule)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.employee = employee
            obj.save()
            messages.success(request, 'График работы сохранён.')
            return redirect('hr:edit_employee', pk=pk)
        messages.error(request, 'Проверьте поля графика работы.')
        context = {
            'form': EmployeeForm(instance=employee),
            'employee': employee,
            'schedule_form': form,
            'positions_by_department_url': reverse('hr:positions_by_department'),
        }
        return render(request, 'site/hr/edit_employee.html', context)

    return redirect('hr:edit_employee', pk=pk)


@need_hr_directory
def delete_employee(request, pk):
    is_hr, _, _ = get_registry_access(request.user)
    if not is_hr:
        return HttpResponseForbidden()

    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        name = employee.user.get_name if employee.user else str(employee.pk)
        user = employee.user
        employee.delete()
        if user:
            user.delete()
        messages.success(request, f'Сотрудник «{name}» удалён.')
        return redirect('hr:employees')
    return redirect('hr:edit_employee', pk=pk)


@need_permission(PermissionEnums.HR)
def calendar(request, category):

    items = (
        CalendarItem.objects.filter(category=category)
        .select_related('user')
        .order_by('-start_date', '-id')
    )

    context = {
        'category': category,
        'category_title': CalendarItemType.get_title(category),
        'items': items,
    }

    return render(request, 'site/hr/calendar.html', context)


@need_permission(PermissionEnums.HR)
def calendar_json(request, category):
    qs = CalendarItem.objects.filter(category=category)
    res = CalendarItemSerializer(qs, many=True)

    return JsonResponse(res.data, safe=False)


@need_permission(PermissionEnums.HR)
def calendar_timeline(request, category):
    """Данные для полосного календаря (командировки и др.)."""
    from django.urls import reverse

    qs = CalendarItem.objects.filter(category=category).select_related('user')
    data = []
    for item in qs:
        user = item.user
        name = user.get_name if hasattr(user, 'get_name') else (
            f'{user.first_name or ""} {user.last_name or ""}'.strip() or user.username
        )
        data.append({
            'id': item.id,
            'content': name,
            'start': item.start_date.isoformat(),
            'end': item.end_date.isoformat(),
            'group': item.title or CalendarItemType.get_title(category),
            'status': 'approved',
            'url': reverse('hr:edit_calendar', args=[item.pk]) + f'?category={category}',
        })
    return JsonResponse(data, safe=False)


@need_permission(PermissionEnums.HR)
def work_calendar_json(request):
    """Производственный календарь для подсветки дней в HR-календаре."""
    start_raw = request.GET.get('start')
    end_raw = request.GET.get('end')
    today = date.today()
    if start_raw and end_raw:
        start = parse_date(start_raw[:10]) or today.replace(day=1)
        end = parse_date(end_raw[:10]) or today
    else:
        start = today.replace(day=1)
        if start.month == 12:
            end = date(start.year + 1, 1, 1) - timedelta(days=1)
        else:
            end = date(start.year, start.month + 1, 1) - timedelta(days=1)

    days = WorkCalendar.objects.filter(date__gte=start, date__lte=end)
    payload = {d.date.isoformat(): d.day_type for d in days}
    return JsonResponse(payload)


@need_permission(PermissionEnums.HR)
def edit_calendar_item(request, pk):
    current = get_or_none(CalendarItem, id=pk) if pk else None
    category = request.GET.get('category') or (current.category if current else CalendarItemType.SECONDMENT.value[0])
    valid_categories = {c[0] for c in CalendarItemType.list()}
    if category not in valid_categories:
        category = CalendarItemType.SECONDMENT.value[0]

    form = CalendarItemForm(
        data=request.POST or None,
        instance=current,
        category=category,
    )

    if request.method == 'POST':
        if form.is_valid():
            new = form.save()
            messages.success(request, 'Командировка сохранена')
            return redirect('hr:calendar', category=new.category)

    is_edit = current is not None
    context = {
        'form': form,
        'category': category,
        'category_title': CalendarItemType.get_title(category),
        'is_edit': is_edit,
        'back_url': reverse('hr:calendar', args=[category]),
    }

    return render(request, 'site/hr/edit_calendar.html', context)


@need_permission(PermissionEnums.HR)
def delete_calendar_item(request, pk):
    current = get_or_none(CalendarItem, id=pk)
    category = current.category

    current.delete()

    return redirect('hr:calendar', category=category)


@need_permission(PermissionEnums.HR_COMPANIES)
def companies(request):
    queryset = Company.objects.all().order_by('name')
    for company in queryset:
        company.employee_count = company.get_employees_count()
    
    context = {
        'companies': queryset,
    }
    return render(request, 'site/hr/companies.html', context)


@need_permission(PermissionEnums.HR_COMPANIES)
def create_company(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        bin_number = request.POST.get('bin_number', '').strip()
        address = request.POST.get('address', '').strip()
        phone = request.POST.get('phone', '').strip()
        email = request.POST.get('email', '').strip()

        if name and bin_number:
            Company.objects.create(
                name=name,
                bin_number=bin_number,
                address=address or None,
                phone=phone or None,
                email=email or None,
            )
            messages.success(request, f'Компания «{name}» добавлена.')
        else:
            messages.error(request, 'Название и БИН обязательны.')

    return redirect('hr:companies')


@need_permission(PermissionEnums.HR_COMPANIES)
def delete_company(request, pk):
    company = get_object_or_404(Company, pk=pk)

    if request.method == 'POST':
        emp_count = company.get_employees_count()
        if emp_count > 0:
            messages.error(
                request,
                f'Нельзя удалить «{company.name}»: к компании привязано {emp_count} '
                f'{"сотрудник" if emp_count == 1 else "сотрудников"}.',
            )
        else:
            dept_count = Department.objects.filter(company=company).count()
            name = company.name
            company.delete()
            if dept_count:
                messages.success(
                    request,
                    f'Компания «{name}» и связанные отделы ({dept_count}) удалены.',
                )
            else:
                messages.success(request, f'Компания «{name}» удалена.')

    return redirect('hr:companies')


@need_permission(PermissionEnums.HR_COMPANIES)
def positions(request):
    queryset = Position.objects.all().order_by('department__name', 'title')
    context = {
        'positions': queryset,
        'departments': Department.objects.all().order_by('name'),
    }
    return render(request, 'site/hr/positions.html', context)


@need_permission(PermissionEnums.HR_COMPANIES)
def create_position(request):
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        department_id = request.POST.get('department')
        description = request.POST.get('description', '').strip()

        if title and department_id:
            try:
                dept = Department.objects.get(pk=department_id)
                Position.objects.create(
                    title=title,
                    department=dept,
                    description=description or None,
                )
                messages.success(request, f'Должность «{title}» добавлена.')
            except Department.DoesNotExist:
                messages.error(request, 'Выбранный отдел не найден.')
        else:
            messages.error(request, 'Название и отдел обязательны.')

    return redirect('hr:positions')


@need_permission(PermissionEnums.HR_COMPANIES)
def departments(request):
    queryset = Department.objects.select_related('company', 'parent').order_by('name')
    context = {
        'departments': queryset,
        'companies': Company.objects.all().order_by('name'),
    }
    return render(request, 'site/hr/departments.html', context)


@need_permission(PermissionEnums.HR_COMPANIES)
def create_department(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        company_id = request.POST.get('company')
        parent_id = request.POST.get('parent') or None
        level_type = request.POST.get('level_type') or 'department'

        if name and company_id:
            try:
                company = Company.objects.get(pk=company_id)
                parent = None
                if parent_id:
                    parent = Department.objects.get(pk=parent_id, company=company)
                Department.objects.create(
                    name=name,
                    company=company,
                    parent=parent,
                    level_type=level_type,
                )
                messages.success(request, f'Отдел «{name}» добавлен.')
            except (Company.DoesNotExist, Department.DoesNotExist):
                messages.error(request, 'Компания или родительский отдел не найдены.')
        else:
            messages.error(request, 'Название и компания обязательны.')

    return redirect('hr:departments')


@need_hr_directory
def positions_by_department(request):
    from django.http import JsonResponse

    dept_id = request.GET.get('department')
    if not dept_id:
        return JsonResponse({'positions': []})

    positions_qs = Position.objects.filter(department_id=dept_id).order_by('title')
    return JsonResponse({
        'positions': [{'id': p.pk, 'title': p.title} for p in positions_qs],
    })


def _is_manager(user):
    """Согласование отпусков: руководитель отдела, HR-менеджер, администратор, журнал посещаемости."""
    employee = getattr(user, 'employee_info', None)
    if employee and employee.head:
        return True

    role = user.role
    if hasattr(role, 'value'):
        role = role.value

    if role in (RoleEnums.HR.value, RoleEnums.ADMINISTRATOR.value):
        return True

    if RolePermissions.checkPermission(role, PermissionEnums.HR_JOURNAL):
        return True

    return False


@need_permission(PermissionEnums.HR_SELF)
def leave_timeline_page(request):
    """HTML-страница Календаря отпусков (данные грузятся через leave_timeline JSON)."""
    return render(request, 'site/hr/leave_timeline.html', {})


@need_permission(PermissionEnums.HR_SELF)
def leave_list(request):
    is_manager = _is_manager(request.user)
    employee = getattr(request.user, 'employee_info', None)

    queryset = LeaveRequest.objects.all().select_related(
        'employee__user',
        'employee__department',
        'leave_type',
        'approver__user',
    ).order_by('-id')

    # Обычный сотрудник видит только свои заявки
    if not is_manager and employee:
        queryset = queryset.filter(employee=employee)

    filter_form = LeaveFilterForm(request.GET)

    if filter_form.is_valid():
        data = filter_form.cleaned_data

        search = data.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(employee__user__first_name__icontains=search) |
                Q(employee__user__last_name__icontains=search)
            )

        if data.get('department'):
            queryset = queryset.filter(employee__department=data['department'])

        if data.get('status'):
            queryset = queryset.filter(status=data['status'])

        if data.get('leave_type'):
            queryset = queryset.filter(leave_type=data['leave_type'])

        if data.get('date_from'):
            queryset = queryset.filter(start_date__gte=data['date_from'])

        if data.get('date_to'):
            queryset = queryset.filter(end_date__lte=data['date_to'])

    page = page_from_request(request)
    paginator = CustomPaginator(queryset, page)

    context = {
        'paginator': paginator,
        'filter_form': filter_form,
        'is_manager': is_manager,
    }

    return render(request, 'site/hr/leave_list.html', context)

@need_permission(PermissionEnums.HR_SELF)
def leave_create(request):
    employee = getattr(request.user, 'employee_info', None)

    if not employee:
        messages.error(request, "Профиль сотрудника не найден.")
        return redirect('hr:leave_list')

    if request.method == 'POST':
        form = LeaveRequestForm(request.POST)
        if form.is_valid():
            leave = form.save(commit=False)
            leave.employee = employee
            leave.status = LeaveStatusEnum.PENDING
            leave.save()
            messages.success(request, "Заявка успешно отправлена.")
            return redirect('hr:leave_list')
    else:
        form = LeaveRequestForm()

    context = {
        'form': form,
    }

    return render(request, 'site/hr/leave_create.html', context)

@need_permission(PermissionEnums.HR_SELF)
def leave_detail(request, pk):
    leave = get_object_or_404(
        LeaveRequest.objects.select_related(
            'employee__user',
            'employee__department',
            'leave_type',
            'approver__user',
        ),
        pk=pk
    )

    is_owner = (leave.employee == getattr(request.user, 'employee_info', None))
    if not is_owner and not _is_manager(request.user):
        messages.error(request, "Нет доступа.")
        return redirect('hr:leave_list')

    context = {
        'leave': leave,
        'is_manager': _is_manager(request.user),
        'is_owner': is_owner,
    }
    return render(request, 'site/hr/leave_detail.html', context)


@need_permission(PermissionEnums.HR_SELF)
def leave_confirm(request, pk):
    """Первый шаг согласования: подтверждение перед финальным одобрением."""
    leave = get_object_or_404(LeaveRequest, pk=pk)

    if request.method == 'POST':
        if _is_manager(request.user) and leave.status == LeaveStatusEnum.PENDING:
            leave.status = LeaveStatusEnum.CONFIRMED
            leave.save()
            messages.success(request, "Заявка подтверждена. Можно одобрить окончательно.")

    return redirect(request.POST.get('next') or 'hr:leave_list')


@need_permission(PermissionEnums.HR_SELF)
def leave_approve(request, pk):
    leave = get_object_or_404(LeaveRequest, pk=pk)

    if request.method == 'POST':
        if _is_manager(request.user) and leave.status == LeaveStatusEnum.CONFIRMED:
            leave.status = LeaveStatusEnum.APPROVED
            leave.approver = request.user.employee_info
            leave.save()
            messages.success(request, "Заявка одобрена.")

    return redirect(request.POST.get('next') or 'hr:leave_list')


@need_permission(PermissionEnums.HR_SELF)
def leave_reject(request, pk):
    leave = get_object_or_404(LeaveRequest, pk=pk)

    if request.method == 'POST':
        if _is_manager(request.user) and leave.status in (
            LeaveStatusEnum.PENDING,
            LeaveStatusEnum.CONFIRMED,
        ):
            leave.status = LeaveStatusEnum.REJECTED
            leave.save()
            messages.warning(request, "Заявка отклонена.")

    return redirect(request.POST.get('next') or 'hr:leave_list')

@need_permission(PermissionEnums.HR_SELF)
def leave_cancel(request, pk):
    leave = get_object_or_404(LeaveRequest, pk=pk, employee__user=request.user)

    if request.method == 'POST':
        if leave.status in [LeaveStatusEnum.DRAFT, LeaveStatusEnum.PENDING]:
            leave.delete()
            messages.success(request, "Заявка успешно удалена.")
    
    return redirect('hr:leave_list')

@need_permission(PermissionEnums.HR_SELF)
def ajax_calculate_days(request):
    start = request.GET.get('start')
    end = request.GET.get('end')

    if not start or not end:
        return JsonResponse({'days': 0})

    try:
        employee = getattr(request.user, 'employee_info', None)
        if not employee:
            return JsonResponse({'days': 0, 'error': 'Профиль сотрудника не найден'})

        from datetime import datetime
        start_date = datetime.strptime(start, '%Y-%m-%d').date()
        end_date = datetime.strptime(end, '%Y-%m-%d').date()

        if start_date > end_date:
            return JsonResponse({'days': 0, 'error': 'Некорректный диапазон дат'})

        temp_leave = LeaveRequest(
            start_date=start_date,
            end_date=end_date,
            employee=employee
        )
        days = temp_leave.calculate_working_days()
        return JsonResponse({'days': days})

    except ValueError:
        return JsonResponse({'days': 0, 'error': 'Некорректный формат даты'})
    except Exception as e:
        return JsonResponse({'days': 0, 'error': str(e)})

@need_permission(PermissionEnums.HR_SELF)
def leave_timeline(request):
    is_manager = _is_manager(request.user)
    curr_employee = getattr(request.user, 'employee_info', None)

    queryset = LeaveRequest.objects.all().select_related(
        'employee__user', 'employee__department__company', 'leave_type'
    )

    # Обычный сотрудник видит только свои заявки
    if not is_manager and curr_employee:
        queryset = queryset.filter(employee=curr_employee)

    company_id = request.GET.get('company')
    department_id = request.GET.get('department')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if company_id:
        queryset = queryset.filter(employee__department__company_id=company_id)
    if department_id:
        queryset = queryset.filter(employee__department_id=department_id)
    if start_date_str:
        queryset = queryset.filter(end_date__gte=parse_date(start_date_str))
    if end_date_str:
        queryset = queryset.filter(start_date__lte=parse_date(end_date_str))

    data = []
    for leave in queryset:
        user = leave.employee.user
        name = f"{user.first_name or ''} {user.last_name or ''}".strip() or user.username
        
        data.append({
            'id': leave.id,
            'content': f"{name} ({leave.leave_type.name})",
            'start': leave.start_date.isoformat(),
            'end': leave.end_date.isoformat(),
            'group': leave.employee.department.name if leave.employee.department else 'Без отдела',
            'status': leave.status,
            'className': f"leave-status-{leave.status}" 
        })

    return JsonResponse(data, safe=False)


@need_permission(PermissionEnums.HR_JOURNAL)
def leave_export_excel(request):
    queryset = LeaveRequest.objects.all().select_related(
        'employee__user', 'employee__department', 'leave_type'
    ).order_by('-start_date')

    company_id = request.GET.get('company')
    department_id = request.GET.get('department')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if company_id:
        queryset = queryset.filter(employee__department__company_id=company_id)
    if department_id:
        queryset = queryset.filter(employee__department_id=department_id)
    if start_date_str:
        queryset = queryset.filter(end_date__gte=parse_date(start_date_str))
    if end_date_str:
        queryset = queryset.filter(start_date__lte=parse_date(end_date_str))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отпуска"

    headers = ["ФИО сотрудника", "Отдел", "Тип отпуска", "Начало", "Конец", "Дней", "Статус"]
    ws.append(headers)

    for leave in queryset:
        user = leave.employee.user
        name = f"{user.first_name or ''} {user.last_name or ''}".strip() or user.username
        dept = leave.employee.department.name if leave.employee.department else "-"
        status_display = dict(LeaveStatusEnum.choices).get(leave.status, leave.status)

        ws.append([
            name,
            dept,
            leave.leave_type.name,
            leave.start_date.strftime("%d.%m.%Y"),
            leave.end_date.strftime("%d.%m.%Y"),
            leave.working_days_count,
            status_display
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="leaves_export.xlsx"'
    
    wb.save(response)
    return response


@need_permission(PermissionEnums.HR)
def vacations(request):
    queryset = Vacation.objects.select_related('employee', 'employee__user').order_by('-start_date', '-id')

    employee_id = request.GET.get('employee')
    if employee_id:
        queryset = queryset.filter(employee_id=employee_id)

    context = {
        'vacations': queryset,
        'selected_employee': employee_id,
    }
    return render(request, 'site/hr/vacations.html', context)


@need_permission(PermissionEnums.HR)
def sick_leaves(request):
    queryset = SickLeave.objects.select_related('employee', 'employee__user').order_by('-start_date', '-id')

    employee_id = request.GET.get('employee')
    if employee_id:
        queryset = queryset.filter(employee_id=employee_id)

    context = {
        'sick_leaves': queryset,
        'selected_employee': employee_id,
    }
    return render(request, 'site/hr/sick_leaves.html', context)


@need_permission(PermissionEnums.HR)
def contracts(request):
    queryset = EmploymentContract.objects.select_related('employee', 'employee__user').order_by('-date', '-id')

    employee_id = request.GET.get('employee')
    if employee_id:
        queryset = queryset.filter(employee_id=employee_id)

    context = {
        'contracts': queryset,
        'selected_employee': employee_id,
    }
    return render(request, 'site/hr/contracts.html', context)

def _round_geo_coord(value):
    """Округление координат до 7 знаков после запятой (лимит DecimalField)."""
    if value is None or value == '':
        return None
    try:
        d = Decimal(str(value))
        return d.quantize(Decimal('0.0000001'), rounding=ROUND_HALF_UP)
    except Exception:
        return None


@login_required
def attendance_checkin(request):
    if request.method == 'GET':
        import pytz
        LOCAL_TZ = pytz.timezone('Asia/Almaty')
        employee = getattr(request.user, 'employee_info', None)
        today = date.today()

        today_marks = []
        completed_types = set()
        next_event = CheckInEnum.DAY_START
        all_done = False

        if employee:
            summary = AttendanceRecord.get_daily_summary(employee, today)
            events = summary.get('details', {})
            order = [
                (CheckInEnum.DAY_START, 'Приход'),
                (CheckInEnum.DAY_END, 'Уход'),
            ]
            for key, label in order:
                ts = events.get(key)
                if ts:
                    completed_types.add(key)
                    local_ts = ts.astimezone(LOCAL_TZ)
                    today_marks.append({
                        'type': key,
                        'label': label,
                        'time': local_ts.strftime('%H:%M'),
                    })

            if CheckInEnum.DAY_START not in completed_types:
                next_event = CheckInEnum.DAY_START
            elif CheckInEnum.DAY_END not in completed_types:
                next_event = CheckInEnum.DAY_END
            else:
                all_done = True
                next_event = None

        event_labels = {
            CheckInEnum.DAY_START: 'Приход',
            CheckInEnum.DAY_END: 'Уход',
        }
        preselect = request.GET.get('event') or (next_event if next_event else CheckInEnum.DAY_START)
        if preselect in completed_types and next_event:
            preselect = next_event

        return render(request, 'site/hr/attendance/checkin.html', {
            'today_marks': today_marks,
            'completed_types': completed_types,
            'next_event': next_event,
            'next_event_label': event_labels.get(next_event, '') if next_event else '',
            'preselect_event': preselect,
            'all_done_today': all_done,
        })


    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не поддерживается. Используйте POST.'}, status=405)

    try:
        data = json.loads(request.body)
        event_type = data.get('event_type')
        photo_base64 = data.get('photo')
        
        ip_address = data.get('ip_address') or request.META.get('REMOTE_ADDR')

        employee = getattr(request.user, 'employee_info', None)
        if not employee:
            return JsonResponse({'error': 'Профиль сотрудника не найден'}, status=403)

        if not event_type or not photo_base64:
            return JsonResponse({'error': 'Не переданы обязательные параметры (event_type, photo)'}, status=400)

        latitude  = _round_geo_coord(data.get('latitude'))
        longitude = _round_geo_coord(data.get('longitude'))

        if ';base64,' in photo_base64:
            format_str, imgstr = photo_base64.split(';base64,')
            ext = format_str.split('/')[-1]
        else:
            imgstr = photo_base64
            ext = 'jpg'
        photo_name = f"checkin_{employee.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
        photo_file = ContentFile(base64.b64decode(imgstr), name=photo_name)

        create_attendance_checkin(
            employee=employee,
            event_type=event_type,
            photo_file=photo_file,
            latitude=latitude,
            longitude=longitude,
            ip_address=ip_address,
        )
        return JsonResponse({'success': True, 'message': 'Отметка успешно сохранена'})

    except ValidationError as e:
        return JsonResponse({'error': e.messages[0]}, status=400)
    except Exception as e:
        return JsonResponse({'error': f"Внутренняя ошибка сервера: {str(e)}"}, status=500)


@login_required
def attendance_resolve_address(request):
    """AJAX: координаты → адрес (для старых отметок без сохранённого адреса)."""
    lat = request.GET.get('lat')
    lng = request.GET.get('lng')
    record_id = request.GET.get('record_id')

    if not lat or not lng:
        return JsonResponse({'address': ''})

    from .geocoding import reverse_geocode
    address = reverse_geocode(lat, lng)
    if address and record_id:
        AttendanceRecord.objects.filter(pk=record_id).update(location_address=address)

    return JsonResponse({'address': address})


@login_required
def attendance_journal(request):
    import pytz
    LOCAL_TZ = pytz.timezone('Asia/Almaty')

    user = request.user
    employee = getattr(user, 'employee_info', None)
    
    is_hr = RolePermissions.checkPermission(user.role, PermissionEnums.HR_JOURNAL)
    is_head = employee and employee.head

    if not is_hr and not is_head:
        return redirect('hr:attendance_my')

    target_date_str = request.GET.get('date', date.today().isoformat())
    target_date = parse_date(target_date_str) or date.today()
    
    employees_qs = Employee.objects.filter(status='active').select_related('user', 'department', 'work_schedule')
    
    if not is_hr and is_head:
        employees_qs = employees_qs.filter(department=employee.department)
    
    department_id = request.GET.get('department')
    if department_id:
        employees_qs = employees_qs.filter(department_id=department_id)

    # Предзагружаем записи за дату: фото + гео + адрес (из БД)
    record_map = {}
    for rec in AttendanceRecord.objects.filter(timestamp__date=target_date).select_related('employee', 'manual_author', 'manual_reason'):
        try:
            photo_url = rec.photo.url if (rec.photo and rec.photo.name) else ''
        except Exception:
            photo_url = ''
        record_map.setdefault(rec.employee_id, {})[rec.event_type] = {
            'photo': photo_url,
            'lat': str(rec.latitude) if rec.latitude else '',
            'lng': str(rec.longitude) if rec.longitude else '',
            'address': rec.location_address or '',
            'is_manual': rec.is_manual,
            'manual_author': rec.manual_author.get_name if rec.manual_author else None,
            'manual_reason': rec.manual_reason.label if rec.manual_reason else None,
        }

    journal = []
    for emp in employees_qs:
        summary = AttendanceRecord.get_daily_summary(emp, target_date)
        events = summary.get('details', {})
        has_records = len(events) > 0

        late = False
        early_leave = False
        start_dt = None
        end_dt = None
        lunch_start_dt = None
        lunch_end_dt   = None

        if has_records:
            start_dt = events.get(CheckInEnum.DAY_START)
            if start_dt:
                local_start = start_dt.astimezone(LOCAL_TZ)
                start_dt = local_start
                late = work_schedule_helper.is_late(emp, local_start)

            end_dt = events.get(CheckInEnum.DAY_END)
            if end_dt:
                local_end = end_dt.astimezone(LOCAL_TZ)
                end_dt = local_end
                if local_end.hour < 18:
                    early_leave = True

            ls = events.get(CheckInEnum.LUNCH_START)
            if ls:
                lunch_start_dt = ls.astimezone(LOCAL_TZ)
            le = events.get(CheckInEnum.LUNCH_END)
            if le:
                lunch_end_dt = le.astimezone(LOCAL_TZ)

        total_work = summary.get('total_work_time', timedelta(0))
        total_hours = total_work.total_seconds() / 3600 if total_work else 0

        emp_recs = record_map.get(emp.id, {})

        def _r(key, field):
            return emp_recs.get(key, {}).get(field, '')

        journal.append({
            'employee':          emp,
            'day_start':         start_dt,
            'day_end':           end_dt,
            'total_hours':       total_hours,
            'late':              late,
            'early_leave':       early_leave,
            'no_record':         not has_records,
            # Приход
            'arrival_photo':     _r(CheckInEnum.DAY_START,   'photo'),
            'arrival_lat':       _r(CheckInEnum.DAY_START,   'lat'),
            'arrival_lng':       _r(CheckInEnum.DAY_START,   'lng'),
            'arrival_address':   _r(CheckInEnum.DAY_START,   'address'),
            # Уход
            'departure_photo':   _r(CheckInEnum.DAY_END,     'photo'),
            'departure_lat':     _r(CheckInEnum.DAY_END,     'lat'),
            'departure_lng':     _r(CheckInEnum.DAY_END,     'lng'),
            'departure_address': _r(CheckInEnum.DAY_END,     'address'),
            # Ручные отметки
            'arrival_is_manual':    _r(CheckInEnum.DAY_START, 'is_manual'),
            'arrival_manual_author': _r(CheckInEnum.DAY_START, 'manual_author'),
            'arrival_manual_reason': _r(CheckInEnum.DAY_START, 'manual_reason'),
            'departure_is_manual':    _r(CheckInEnum.DAY_END, 'is_manual'),
            'departure_manual_author': _r(CheckInEnum.DAY_END, 'manual_author'),
            'departure_manual_reason': _r(CheckInEnum.DAY_END, 'manual_reason'),
        })

    departments = Department.objects.all() if is_hr else [employee.department]

    from django.conf import settings as django_settings
    lunch_enabled = not (
        CheckInEnum.LUNCH_START in getattr(django_settings, 'ATTENDANCE_DISABLED_TYPES', []) and
        CheckInEnum.LUNCH_END in getattr(django_settings, 'ATTENDANCE_DISABLED_TYPES', [])
    )

    return render(request, 'site/hr/attendance_journal.html', {
        'journal': journal,
        'target_date': target_date,
        'departments': departments,
        'is_hr': is_hr,
        'lunch_enabled': lunch_enabled,
    })


def _attendance_registry_required(view):
    """
    Тот же круг доступа, что у AttendanceRegistryPermission на DRF-стороне
    (hr/api.py::AttendanceRegistryViewSet) — держим оба места в синхроне
    вручную, т.к. общего PermissionEnum под этот конкретный набор ролей нет.
    """
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            response = redirect('account:auth')
            response['Location'] += f"?next={request.path}"
            return response

        role = request.user.role
        if hasattr(role, 'value'):
            role = role.value
        employee = getattr(request.user, 'employee_info', None)
        allowed = (
            request.user.is_superuser
            or role in (RoleEnums.ADMINISTRATOR.value, RoleEnums.HR.value, RoleEnums.OWNER.value)
            or (employee is not None and getattr(employee, 'head', False))
        )
        if not allowed:
            raise PermissionDenied('Реестр посещаемости')
        return view(request, *args, **kwargs)
    return wrapper


@_attendance_registry_required
def attendance_registry(request):
    user = request.user
    role = user.role
    if hasattr(role, 'value'):
        role = role.value
    employee = getattr(user, 'employee_info', None)
    is_full_access = user.is_superuser or role in (
        RoleEnums.ADMINISTRATOR.value, RoleEnums.HR.value, RoleEnums.OWNER.value,
    )

    if is_full_access:
        departments = Department.objects.all()
        employees = Employee.objects.filter(status='active').select_related('user', 'department')
    else:
        dept_ids = list(
            employee.department.get_descendants(include_self=True).values_list('id', flat=True)
        )
        departments = Department.objects.filter(id__in=dept_ids)
        employees = Employee.objects.filter(
            department_id__in=dept_ids, status='active',
        ).select_related('user', 'department')

    return render(request, 'site/hr/attendance_registry.html', {
        'departments': departments,
        'employees': employees,
    })


@login_required
def attendance_my(request):
    import pytz
    LOCAL_TZ = pytz.timezone('Asia/Almaty')

    curr_employee = getattr(request.user, 'employee_info', None)
    
    target_emp_id = request.GET.get('employee_id')
    if target_emp_id:
        is_hr = RolePermissions.checkPermission(request.user.role, PermissionEnums.HR_JOURNAL)
        is_head = curr_employee and curr_employee.head
        
        if is_hr:
            employee = get_object_or_404(Employee, id=target_emp_id)
        elif is_head:
            employee = get_object_or_404(Employee, id=target_emp_id, department=curr_employee.department)
        else:
            return redirect('hr:attendance_my')
    else:
        employee = curr_employee

    if not employee:
        return redirect('dashboard:dashboard')

    try:
        view_month = int(request.GET.get('month', date.today().month))
        view_year = int(request.GET.get('year', date.today().year))
    except ValueError:
        view_month = date.today().month
        view_year = date.today().year

    import calendar as calendar_module
    _, num_days = calendar_module.monthrange(view_year, view_month)

    # Предзагружаем все записи за месяц: фото + гео
    month_start = date(view_year, view_month, 1)
    month_end   = date(view_year, view_month, num_days)
    my_records = {}  # date -> {event_type: {photo, lat, lng, address}}
    for rec in AttendanceRecord.objects.filter(
        employee=employee, timestamp__date__range=(month_start, month_end)
    ):
        try:
            photo_url = rec.photo.url if (rec.photo and rec.photo.name) else ''
        except Exception:
            photo_url = ''
        d = rec.timestamp.date()
        my_records.setdefault(d, {})[rec.event_type] = {
            'photo': photo_url,
            'lat': str(rec.latitude) if rec.latitude else '',
            'lng': str(rec.longitude) if rec.longitude else '',
            'address': rec.location_address or '',
        }

    attendance_list = []
    for day in range(num_days, 0, -1):
        current_day = date(view_year, view_month, day)
        if current_day > date.today():
            continue

        summary = AttendanceRecord.get_daily_summary(employee, current_day)
        events = summary.get('details', {})

        late = False
        early_leave = False

        start_dt = events.get(CheckInEnum.DAY_START)
        if start_dt:
            start_dt = start_dt.astimezone(LOCAL_TZ)
            late = work_schedule_helper.is_late(employee, start_dt)

        end_dt = events.get(CheckInEnum.DAY_END)
        if end_dt:
            end_dt = end_dt.astimezone(LOCAL_TZ)
            if end_dt.hour < 18:
                early_leave = True

        lunch_start = events.get(CheckInEnum.LUNCH_START)
        if lunch_start:
            lunch_start = lunch_start.astimezone(LOCAL_TZ)
        lunch_end = events.get(CheckInEnum.LUNCH_END)
        if lunch_end:
            lunch_end = lunch_end.astimezone(LOCAL_TZ)

        total_hours = summary.get('total_work_time', timedelta(0)).total_seconds() / 3600
        day_recs = my_records.get(current_day, {})

        def _dr(key, field):
            return day_recs.get(key, {}).get(field, '')

        attendance_list.append({
            'date':              current_day,
            'day_start':         start_dt,
            'day_end':           end_dt,
            'lunch_start':       lunch_start,
            'lunch_end':         lunch_end,
            'total_hours':       total_hours,
            'late':              late,
            'early_leave':       early_leave,
            'no_record':         len(events) == 0,
            'arrival_photo':     _dr(CheckInEnum.DAY_START,   'photo'),
            'arrival_lat':       _dr(CheckInEnum.DAY_START,   'lat'),
            'arrival_lng':       _dr(CheckInEnum.DAY_START,   'lng'),
            'arrival_address':   _dr(CheckInEnum.DAY_START,   'address'),
            'lunch_start_photo': _dr(CheckInEnum.LUNCH_START, 'photo'),
            'lunch_end_photo':   _dr(CheckInEnum.LUNCH_END,   'photo'),
            'departure_photo':   _dr(CheckInEnum.DAY_END,     'photo'),
            'departure_lat':     _dr(CheckInEnum.DAY_END,     'lat'),
            'departure_lng':     _dr(CheckInEnum.DAY_END,     'lng'),
            'departure_address': _dr(CheckInEnum.DAY_END,     'address'),
        })

    prev_month_date = date(view_year, view_month, 1) - timedelta(days=1)
    next_month_date = date(view_year, view_month, 28) + timedelta(days=5)

    from django.conf import settings as django_settings
    lunch_enabled = not (
        CheckInEnum.LUNCH_START in getattr(django_settings, 'ATTENDANCE_DISABLED_TYPES', []) and
        CheckInEnum.LUNCH_END in getattr(django_settings, 'ATTENDANCE_DISABLED_TYPES', [])
    )


    return render(request, 'site/hr/attendance_my.html', {
        'attendance_list': attendance_list,
        'view_date': date(view_year, view_month, 1),
        'prev_month': prev_month_date,
        'next_month': next_month_date if next_month_date <= date.today() else None,
        'employee': employee,
        'is_own_profile': employee == curr_employee,
        'lunch_enabled': lunch_enabled
    })


@need_permission(PermissionEnums.HR_SELF)
def documents_list(request):
    is_hr, is_head, curr_employee = get_registry_access(request.user)

    employee_id = request.GET.get('employee_id')

    if not is_hr and not is_head and curr_employee and not employee_id:
        return redirect(f"{reverse('hr:documents_list')}?employee_id={curr_employee.pk}")


    if employee_id:
        selected_employee = get_object_or_404(Employee, pk=employee_id)
    
        if not is_hr and not is_head:
            if curr_employee is None or curr_employee.pk != selected_employee.pk:
                return HttpResponseForbidden()
        elif is_head and not is_hr:
            if curr_employee and selected_employee.department != curr_employee.department:
                return HttpResponseForbidden()
    
        queryset = EmployeeDocument.objects.filter(employee=selected_employee).select_related(
            'employee__user', 'employee__department'
        ).order_by('-created_at')
        page = page_from_request(request)
        doc_paginator = CustomPaginator(queryset, page)
        return render(request, 'site/hr/documents_list.html', {
            'paginator': doc_paginator,
            'selected_employee': selected_employee,
            'is_hr': is_hr or is_head,
            'employee_mode': True,
        })

    employees_qs = (
        Employee.objects
        .filter(status='active')
        .select_related('user', 'department')
        .prefetch_related('documents')
        .order_by('user__last_name', 'user__first_name')
    )
    if not is_hr:
        if is_head and curr_employee:
            employees_qs = employees_qs.filter(department=curr_employee.department)
        elif curr_employee:
            employees_qs = employees_qs.filter(pk=curr_employee.pk)
        else:
            employees_qs = employees_qs.none()

    page = page_from_request(request)
    emp_paginator = CustomPaginator(employees_qs, page)
    return render(request, 'site/hr/documents_list.html', {
        'paginator': emp_paginator,
        'is_hr': is_hr or is_head,
        'employee_mode': False,
    })

@need_hr_registry
def permits_list(request):
    is_hr, is_head, _ = get_registry_access(request.user)
    queryset = EmployeeWorkPermit.objects.select_related('employee__user', 'employee__department', 'category').order_by('expiry_date')
    queryset = filter_by_access(queryset, request.user)
    filter_form = PermitFilterForm(request.GET)
    if filter_form.is_valid():
        data = filter_form.cleaned_data
        if data.get('search'):
            queryset = queryset.filter(
                Q(employee__user__first_name__icontains=data['search']) |
                Q(employee__user__last_name__icontains=data['search'])
            )
        if data.get('department'):
            queryset = queryset.filter(employee__department=data['department'])
        if data.get('category'):
            queryset = queryset.filter(category=data['category'])
        if data.get('expiring_soon'):
            queryset = queryset.filter(expiry_date__lte=date.today() + timedelta(days=30), expiry_date__gte=date.today())
        if data.get('expired'):
            queryset = queryset.filter(expiry_date__lt=date.today())
    page = page_from_request(request)
    paginator = CustomPaginator(queryset, page)
    return render(request, 'site/hr/permits_list.html', {
        'paginator': paginator,
        'filter_form': filter_form,
        'is_hr': is_hr or is_head,
        'date_today': date.today(),
        'expiration_threshold': date.today() + timedelta(days=30),
    })

@need_hr_registry
def documents_create(request):
    is_hr, is_head, _ = get_registry_access(request.user)
    if not is_hr and not is_head:
        return HttpResponseForbidden()
    emp_id = request.GET.get('employee_id')
    initial = {}
    if emp_id:
        initial['employee'] = get_object_or_404(Employee, pk=emp_id)
    form = EmployeeDocumentForm(request.POST or None, request.FILES or None, initial=initial)
    if request.method == 'POST' and form.is_valid():
        doc = form.save()
        messages.success(request, "Документ добавлен.")
        return redirect(f"{reverse('hr:employee_detail', args=[doc.employee_id])}?tab=documents")
    return render(request, 'site/hr/documents_form.html', {'form': form, 'title': 'Добавить документ'})

@need_hr_registry
def documents_edit(request, pk):
    is_hr, is_head, _ = get_registry_access(request.user)
    if not is_hr and not is_head:
        return HttpResponseForbidden()
    doc = get_object_or_404(
        filter_by_access(EmployeeDocument.objects.all(), request.user),
        pk=pk,
    )
    form = EmployeeDocumentForm(request.POST or None, request.FILES or None, instance=doc)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Документ обновлён.")
        return redirect('hr:documents_list')
    return render(request, 'site/hr/documents_form.html', {'form': form, 'title': 'Редактировать документ'})

@need_hr_registry
def documents_delete(request, pk):
    is_hr, is_head, _ = get_registry_access(request.user)
    if not is_hr and not is_head:
        return HttpResponseForbidden()
    doc = get_object_or_404(EmployeeDocument, pk=pk)
    if request.method == 'POST':
        doc.delete()
        messages.success(request, "Документ удалён.")
    return redirect('hr:documents_list')

@need_hr_registry
def documents_export(request):
    is_hr, is_head, _ = get_registry_access(request.user)
    if not is_hr and not is_head:
        return HttpResponseForbidden()
    queryset = EmployeeDocument.objects.select_related('employee__user', 'employee__department').order_by('-created_at')
    queryset = filter_by_access(queryset, request.user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Документы"
    ws.append(["Сотрудник", "Отдел", "Тип", "Название", "Версия", "Статус", "Дата подписания", "Истекает"])
    for doc in queryset:
        user = doc.employee.user
        name = f"{user.first_name or ''} {user.last_name or ''}".strip() or user.username
        ws.append([
            name,
            doc.employee.department.name if doc.employee.department else "-",
            doc.get_doc_type_display(),
            doc.title,
            doc.version,
            doc.get_status_display(),
            doc.signed_at.strftime("%d.%m.%Y") if doc.signed_at else "-",
            doc.expires_at.strftime("%d.%m.%Y") if doc.expires_at else "-",
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="documents_export.xlsx"'
    wb.save(response)
    return response

@need_hr_registry
def documents_esigner_send(request, pk):
    is_hr, is_head, _ = get_registry_access(request.user)
    if not is_hr and not is_head:
        return HttpResponseForbidden()

    doc = get_object_or_404(
        filter_by_access(EmployeeDocument.objects.all(), request.user),
        pk=pk,
    )

    if request.method != 'POST':
        return redirect('hr:documents_list')

    if not (doc.employee.iin or '').isdigit() or len(doc.employee.iin) != 12:
        messages.error(request, "У сотрудника не заполнен корректный ИИН - подписание невозможно.")
        return redirect('hr:documents_list')

    signers = [{"bin_or_iin": doc.employee.iin, "is_company": False}]
    try:
        signing = send_for_signing(doc, "file", signers)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect('hr:documents_list')
    return redirect(signing.sign_url)


@need_hr_registry
def permits_create(request):
    is_hr, is_head, _ = get_registry_access(request.user)
    if not is_hr and not is_head:
        return HttpResponseForbidden()
    emp_id = request.GET.get('employee_id')
    initial = {}
    if emp_id:
        initial['employee'] = get_object_or_404(Employee, pk=emp_id)
    form = EmployeeWorkPermitForm(request.POST or None, request.FILES or None, initial=initial)
    if request.method == 'POST' and form.is_valid():
        permit = form.save()
        messages.success(request, "Допуск добавлен.")
        return redirect(f"{reverse('hr:employee_detail', args=[permit.employee_id])}?tab=permits")
    return render(request, 'site/hr/permits_form.html', {'form': form, 'title': 'Добавить допуск'})

@need_hr_registry
def permits_edit(request, pk):
    is_hr, is_head, _ = get_registry_access(request.user)
    if not is_hr and not is_head:
        return HttpResponseForbidden()
    permit = get_object_or_404(
        filter_by_access(EmployeeWorkPermit.objects.all(), request.user),
        pk=pk,
    )
    form = EmployeeWorkPermitForm(request.POST or None, request.FILES or None, instance=permit)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Допуск обновлён.")
        return redirect('hr:permits_list')
    return render(request, 'site/hr/permits_form.html', {'form': form, 'title': 'Редактировать допуск'})

@need_hr_registry
def permits_delete(request, pk):
    is_hr, is_head, _ = get_registry_access(request.user)
    if not is_hr and not is_head:
        return HttpResponseForbidden()
    permit = get_object_or_404(EmployeeWorkPermit, pk=pk)
    if request.method == 'POST':
        permit.delete()
        messages.success(request, "Допуск удалён.")
    return redirect('hr:permits_list')

@need_hr_registry
def permits_export(request):
    queryset = EmployeeWorkPermit.objects.select_related('employee__user', 'employee__department', 'category').order_by('expiry_date')
    queryset = filter_by_access(queryset, request.user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Допуски"
    ws.append(["Сотрудник", "Отдел", "Категория", "Номер документа", "Дата выдачи", "Истекает", "Статус"])
    for permit in queryset:
        user = permit.employee.user
        name = f"{user.first_name or ''} {user.last_name or ''}".strip() or user.username
        ws.append([
            name,
            permit.employee.department.name if permit.employee.department else "-",
            permit.category.name,
            permit.document_number or "-",
            permit.issue_date.strftime("%d.%m.%Y"),
            permit.expiry_date.strftime("%d.%m.%Y"),
            permit.status_label,
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="permits_export.xlsx"'
    wb.save(response)
    return response

@need_hr_registry
def permits_esigner_send(request, pk):
    is_hr, is_head, _ = get_registry_access(request.user)
    if not is_hr and not is_head:
        return HttpResponseForbidden()

    permit = get_object_or_404(
        filter_by_access(EmployeeWorkPermit.objects.all(), request.user),
        pk=pk,
    )

    if request.method != 'POST':
        return redirect('hr:permits_list')

    if not (permit.employee.iin or '').isdigit() or len(permit.employee.iin) != 12:
        messages.error(request, "У сотрудника не заполнен корректный ИИН - подписание невозможно.")
        return redirect('hr:permits_list')

    signers = [{"bin_or_iin": permit.employee.iin, "is_company": False}]
    try:
        signing = send_for_signing(permit, "scan", signers)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect('hr:permits_list')
    return redirect(signing.sign_url)


@need_hr_registry
def certifications_list(request):
    is_hr, is_head, _ = get_registry_access(request.user)
    queryset = EmployeeCertification.objects.select_related('employee__user', 'employee__department').order_by('expiry_date')
    queryset = filter_by_access(queryset, request.user)
    filter_form = CertificationFilterForm(request.GET)
    if filter_form.is_valid():
        data = filter_form.cleaned_data
        if data.get('search'):
            queryset = queryset.filter(
                Q(employee__user__first_name__icontains=data['search']) |
                Q(employee__user__last_name__icontains=data['search']) |
                Q(cert_type__icontains=data['search'])
            )
        if data.get('department'):
            queryset = queryset.filter(employee__department=data['department'])
        if data.get('cert_type'):
            queryset = queryset.filter(cert_type__icontains=data['cert_type'])
        if data.get('status'):
            queryset = queryset.filter(status=data['status'])
        if data.get('expiring_soon'):
            queryset = queryset.filter(expiry_date__lte=date.today() + timedelta(days=30), expiry_date__gte=date.today())
    page = page_from_request(request)
    paginator = CustomPaginator(queryset, page)
    return render(request, 'site/hr/certifications_list.html', {
        'paginator': paginator,
        'filter_form': filter_form,
        'is_hr': is_hr or is_head,
    })

@need_hr_registry
def certifications_create(request):
    is_hr, is_head, _ = get_registry_access(request.user)
    if not is_hr and not is_head:
        return HttpResponseForbidden()
    emp_id = request.GET.get('employee_id')
    initial = {}
    if emp_id:
        initial['employee'] = get_object_or_404(Employee, pk=emp_id)
    form = EmployeeCertificationForm(request.POST or None, request.FILES or None, initial=initial)
    if request.method == 'POST' and form.is_valid():
        cert = form.save()
        messages.success(request, "Сертификация добавлена.")
        return redirect(f"{reverse('hr:employee_detail', args=[cert.employee_id])}?tab=certifications")
    return render(request, 'site/hr/certifications_form.html', {'form': form, 'title': 'Добавить сертификацию'})

@need_hr_registry
def certifications_edit(request, pk):
    is_hr, is_head, _ = get_registry_access(request.user)
    if not is_hr and not is_head:
        return HttpResponseForbidden()
    cert = get_object_or_404(EmployeeCertification, pk=pk)
    form = EmployeeCertificationForm(request.POST or None, request.FILES or None, instance=cert)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Сертификация обновлена.")
        return redirect('hr:certifications_list')
    return render(request, 'site/hr/certifications_form.html', {'form': form, 'title': 'Редактировать сертификацию'})

@need_hr_registry
def certifications_delete(request, pk):
    is_hr, is_head, _ = get_registry_access(request.user)
    if not is_hr and not is_head:
        return HttpResponseForbidden()
    cert = get_object_or_404(EmployeeCertification, pk=pk)
    if request.method == 'POST':
        cert.delete()
        messages.success(request, "Сертификация удалена.")
    return redirect('hr:certifications_list')

@need_hr_registry
def certifications_export(request):
    queryset = EmployeeCertification.objects.select_related('employee__user', 'employee__department').order_by('expiry_date')
    queryset = filter_by_access(queryset, request.user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сертификации"
    ws.append(["Сотрудник", "Отдел", "Тип", "Номер", "Дата выдачи", "Истекает", "Статус"])
    for cert in queryset:
        user = cert.employee.user
        name = f"{user.first_name or ''} {user.last_name or ''}".strip() or user.username
        ws.append([
            name,
            cert.employee.department.name if cert.employee.department else "-",
            cert.cert_type,
            cert.certificate_number or "-",
            cert.issue_date.strftime("%d.%m.%Y"),
            cert.expiry_date.strftime("%d.%m.%Y") if cert.expiry_date else "-",
            cert.status,
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="certifications_export.xlsx"'
    wb.save(response)
    return response

@need_permission(PermissionEnums.HR_JOURNAL)
def attendance_export(request):
    import pytz
    from django.conf import settings as django_settings
    LOCAL_TZ = pytz.timezone('Asia/Almaty')

    target_date_str = request.GET.get('date', date.today().isoformat())
    target_date = parse_date(target_date_str) or date.today()

    lunch_enabled = not (
        CheckInEnum.LUNCH_START in getattr(django_settings, 'ATTENDANCE_DISABLED_TYPES', []) and
        CheckInEnum.LUNCH_END in getattr(django_settings, 'ATTENDANCE_DISABLED_TYPES', [])
    )

    employees_qs = Employee.objects.filter(status='active').select_related('user', 'department')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Посещаемость"

    headers = ["Сотрудник", "Отдел", "Приход", "Уход"]
    if lunch_enabled:
        headers += ["Начало обеда", "Конец обеда"]
    headers += ["Рабочих часов", "Опоздание", "Ранний уход"]
    ws.append(headers)

    for emp in employees_qs:
        summary = AttendanceRecord.get_daily_summary(emp, target_date)
        events = summary.get('details', {})

        user = emp.user
        name = f"{user.first_name or ''} {user.last_name or ''}".strip() or user.username
        dept = emp.department.name if emp.department else "-"

        start_dt = events.get(CheckInEnum.DAY_START)
        end_dt = events.get(CheckInEnum.DAY_END)
        lunch_start = events.get(CheckInEnum.LUNCH_START)
        lunch_end = events.get(CheckInEnum.LUNCH_END)

        fmt = lambda dt: dt.astimezone(LOCAL_TZ).strftime('%H:%M') if dt else "-"

        total_hours = summary.get('total_work_time', timedelta(0)).total_seconds() / 3600

        late = False
        early_leave = False
        if start_dt:
            late = work_schedule_helper.is_late(emp, start_dt.astimezone(LOCAL_TZ))
        if end_dt:
            early_leave = end_dt.astimezone(LOCAL_TZ).hour < 18

        row = [name, dept, fmt(start_dt), fmt(end_dt)]
        if lunch_enabled:
            row += [fmt(lunch_start), fmt(lunch_end)]
        row += [
            round(total_hours, 2),
            "Да" if late else "Нет",
            "Да" if early_leave else "Нет",
        ]
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="attendance_{target_date}.xlsx"'
    wb.save(response)
    return response

@need_permission(PermissionEnums.HR)
def manual_attendance(request):
    from account.models import Employee
    from .models import AttendanceManualReason

    employees = Employee.objects.filter(status='active').select_related('user').order_by('user__last_name')
    reasons = AttendanceManualReason.objects.filter(is_active=True)

    return render(request, 'site/hr/manual_attendance.html', {
        'employees': employees,
        'reasons': reasons,
    })

@need_permission(PermissionEnums.HR)
def manual_attendance_report(request):
    from account.models import Employee

    employees = Employee.objects.filter(status='active').select_related('user').order_by('user__last_name')

    return render(request, 'site/hr/manual_attendance_report.html', {
        'employees': employees,
    })
@need_permission(PermissionEnums.HR_JOURNAL)
def qr_points_list(request):
    from hr.models import QRPoint
    points = QRPoint.objects.all().order_by('name')
    return render(request, 'site/hr/attendance/qr_points_list.html', {
        'points': points,
    })


@need_permission(PermissionEnums.HR_JOURNAL)
def qr_point_create(request):
    from hr.models import QRPoint
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        location = request.POST.get('location', '').strip()
        if not name:
            return render(request, 'site/hr/attendance/qr_point_form.html', {
                'title': 'Новая QR-точка',
                'error': 'Название обязательно',
                'form_name': name,
                'form_location': location,
            })
        point = QRPoint.objects.create(name=name, location=location, created_by=request.user)
        # Сразу открываем экран киоска — обычно точку создают, чтобы
        # немедленно вывести её на экран у входа.
        return redirect('hr:qr_kiosk', pk=point.pk)
    return render(request, 'site/hr/attendance/qr_point_form.html', {
        'title': 'Новая QR-точка',
        'form_name': '',
        'form_location': '',
    })


@need_permission(PermissionEnums.HR_JOURNAL)
def qr_point_edit(request, pk):
    from hr.models import QRPoint
    point = get_object_or_404(QRPoint, pk=pk)
    if request.method == 'POST':
        point.name = request.POST.get('name', point.name).strip()
        point.location = request.POST.get('location', point.location).strip()
        point.is_active = request.POST.get('is_active') == 'on'
        point.save()
        return redirect('hr:qr_points_list')
    return render(request, 'site/hr/attendance/qr_point_form.html', {
        'title': 'Редактировать QR-точку',
        'point': point,
    })


@need_permission(PermissionEnums.HR_JOURNAL)
def qr_point_delete(request, pk):
    from hr.models import QRPoint
    point = get_object_or_404(QRPoint, pk=pk)
    if request.method == 'POST':
        point.delete()
        return redirect('hr:qr_points_list')
    return render(request, 'site/hr/attendance/qr_point_confirm_delete.html', {'point': point})


@need_permission(PermissionEnums.HR_JOURNAL)
def qr_kiosk(request, pk):
    from hr.models import QRPoint
    point = get_object_or_404(QRPoint, pk=pk, is_active=True)
    return render(request, 'site/hr/attendance/qr_kiosk.html', {
        'point': point,
        'token_url': reverse('hr:qr_kiosk_token', args=[pk]),
        'checkin_types': [
            {'value': 'day_start', 'label': 'Приход'},
            {'value': 'day_end', 'label': 'Уход'},
        ],
    })


def qr_kiosk_token(request, pk):
    from hr.models import QRPoint, QRToken
    point = get_object_or_404(QRPoint, pk=pk, is_active=True)
    event_type = request.GET.get('event_type', 'day_start')

    token_value = secrets.token_urlsafe(32)
    expires_at = timezone.now() + timedelta(seconds=45)

    QRToken.objects.create(
        token=token_value,
        qr_point=point,
        event_type=event_type,
        expires_at=expires_at,
        ip_address=request.META.get('REMOTE_ADDR'),
    )

    scan_url = request.build_absolute_uri(
        reverse('hr:qr_checkin') + f'?token={token_value}'
    )

    return JsonResponse({
        'token': token_value,
        'scan_url': scan_url,
        'expires_at': expires_at.isoformat(),
        'expires_in': 45,
        'event_type': event_type,
    })


@login_required
def qr_checkin(request):
    from hr.models import QRToken, QRScanAudit
    from account.models import Employee

    token_value = request.POST.get('token') or request.GET.get('token', '')
    ip = request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip() or request.META.get('REMOTE_ADDR')

    def _audit(action, token_str, qr_point=None):
        QRScanAudit.objects.create(
            token=token_str,
            qr_point=qr_point,
            user=request.user if request.user.is_authenticated else None,
            action=action,
            ip_address=ip,
        )

    if not token_value:
        return JsonResponse({'success': False, 'error': 'Недействительный QR-код'}, status=400, json_dumps_params={'ensure_ascii': False})

    try:
        qr_token = QRToken.objects.select_related('qr_point').get(token=token_value)
    except QRToken.DoesNotExist:
        _audit(QRScanAudit.ACTION_INVALID, token_value)
        return JsonResponse({'success': False, 'error': 'Недействительный QR-код'}, status=400, json_dumps_params={'ensure_ascii': False})

    if qr_token.is_expired:
        _audit(QRScanAudit.ACTION_EXPIRED, token_value, qr_token.qr_point)
        return JsonResponse({'success': False, 'error': 'QR-код истёк, отсканируйте текущий код'}, status=410, json_dumps_params={'ensure_ascii': False})

    if qr_token.is_used_by(request.user):
        _audit(QRScanAudit.ACTION_REPLAY, token_value, qr_token.qr_point)
        return JsonResponse({'success': False, 'error': 'Этот QR-код уже использован'}, status=409, json_dumps_params={'ensure_ascii': False})

    try:
        employee = request.user.employee_info
    except Exception:
        _audit(QRScanAudit.ACTION_INVALID, token_value, qr_token.qr_point)
        return JsonResponse({'success': False, 'error': 'Профиль сотрудника не найден'}, status=403, json_dumps_params={'ensure_ascii': False})

    from hr.services import create_attendance_checkin

    create_attendance_checkin(
        employee=employee,
        event_type=qr_token.event_type,
        photo_file=None,
        ip_address=ip,
        source='qr',
    )

    qr_token.used_by.add(request.user)
    _audit(QRScanAudit.ACTION_SUCCESS, token_value, qr_token.qr_point)

    return JsonResponse({'success': True, 'message': 'Отметка успешно создана'}, json_dumps_params={'ensure_ascii': False})
